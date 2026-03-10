"""
Generate NCBI SRA and Genome submission templates from BERDL provenance data.

This script walks provenance from genomes to find:
1. Raw reads with FASTQ files on genomics.lbl.gov
2. Samples used to isolate strains
3. Protocols used for sequencing and assembly
4. Location information for biosamples

It generates:
- Biosample table (SRA Microbe 1.0 package)
- SRA metadata table for FASTQ submissions
- Genome metadata table for FASTA contig submissions
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
from collections import defaultdict
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from pathlib import Path
import sys
from openpyxl import load_workbook
import requests

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from tools.walk_provenance import (  # noqa: E402
    build_downstream_lookup,
    discover_tables,
    get_table_schema,
    load_process_cache,
    NameResolver,
    parse_token,
    select_all_rows,
    set_debug,
    walk_provenance,
)
from tools import walk_provenance as walk_provenance_module  # noqa: E402


FASTQ_HOST = "genomics.lbl.gov"
DEFAULT_BASE_URL = "https://hub.berdl.kbase.us/apis/mcp"
DEFAULT_OUTPUT_DIR = "ncbi_submission"
DEFAULT_EDR_PATH = "/mnt/net/dipa.jmcnet/data/edr"
EDR_URL_PREFIX = "https://genomics.lbl.gov/enigma-data/"
MAX_GENOMES_PER_SUBMISSION = 400
MAX_SRA_ROWS_PER_SUBMISSION = 1000
DEFAULT_REMAINING_UNKNOWN_ASSEMBLIES_PATH = (
    REPO_ROOT / "genome_upload" / "remaining_unknown_assemblies.txt"
)
PROTOCOL_TABLE = "sdt_protocol"
READ_COVERAGE_TABLE = "ddt_brick0000521"
READ_COVERAGE_COLUMN = "read_coverage_statistic_average_count_unit"
GENBANK_LINK_TABLE = "ddt_brick0000529"
Bacteria_AVAILABLE_FROM = (
    "Romy Chakraborty Lab, Berkeley National Lab, Berkeley CA, USA"
)
BIOSAMPLE_COLLECTED_BY_OVERRIDE = "Hazen Lab"
BIOSAMPLE_TEMPLATE_CANDIDATES = [
    REPO_ROOT / "templates" / "Microbe.1.0.xlsx",
    Path("ncbi_submission") / "Microbe.1.0.xlsx",
    Path("ncbi_submission") / "MIcrobe.1.0.xlsx",
    Path("genome_upload") / "ncbi_submission" / "Microbe.1.0.xlsx",
]
SRA_TEMPLATE_CANDIDATES = [
    REPO_ROOT / "templates" / "SRA_metadata.xlsx",
    Path("ncbi_submission") / "SRA_metadata.xlsx",
    Path("genome_upload") / "ncbi_submission" / "SRA_metadata.xlsx",
]
GENOME_TEMPLATE_CANDIDATES = [
    REPO_ROOT / "templates" / "Template_GenomeBatch.xlsx",
    Path("ncbi_submission") / "Template_GenomeBatch.xlsx",
    Path("genome_upload") / "ncbi_submission" / "Template_GenomeBatch.xlsx",
]
SAMPLE_METADATA_CANDIDATES = [
    Path("sample_metadata.tsv"),
    Path("genome_upload") / "sample_metadata.tsv",
]
SRA_DEFAULT_DESIGN_DESCRIPTION = (
    "Whole genome shotgun sequencing for isolate characterization."
)
SRA_ILLUMINA_DESIGN_DESCRIPTION = (
    "Illumina libraries were prepared using ~500ng of input isolate DNA with the "
    "Illumina DNA prep kit  with  IDT(R) for Illumina(R) DNA/RNA sets A-D or with "
    "Illumina Nextera DNA Flex kits with Nextera DNA CD Indexes."
)
SRA_ONT_PLASMIDSAURUS_DESIGN_DESCRIPTION = (
    "Oxford Nanopore libraries were prepared using the Rapid Barcoding Kit 96 V14 "
    "(part # SQK-RBK114.96), Libraries were sequenced on a PromethION P24 with "
    "R10.4.1. flow cell.  Base calling with ont-doradod-for-promethion on "
    "super-accurate mode, minimum Qscore 10, adapters trimmed by MinKnow."
)
SRA_ONT_LAUREN_DESIGN_DESCRIPTION = (
    "The native barcoding expansion (EXP-NBD104; Oxford Nanopore Technologies) and "
    "ligation sequencing (LSK-SQK109; Oxford Nanopore Technologies) kits were used "
    "for barcoding and adapter ligation, respectively.  Libraries were sequenced on "
    "a R9.4.1 flow cell on a MinION device (Oxford Nanopore Technologies).  Base "
    "calling, adapter removal, demultiplexing, and quality filtering were performed "
    "using Guppy v4.0."
)
SRA_PACBIO_SEQUEL_DESIGN_DESCRIPTION = (
    "gDNA from different genera of bacteria were mixed and libraries were "
    "constructed and sequenced with Pacbio SMRTbell Library prep."
)

FLYE_292_B1786_GENOMES = {
    "GW821-FHT02D03",
    "GW821-FHT03E06",
    "GW821-FHT03G07",
    "GW821-FHT09A04",
    "GW821-FHT09A07",
    "GW822-FHT05B01",
    "GW823-FHT04A11",
}

FLYE_29_B1768_GENOMES = {
    "GW821-FHT01A05",
    "GW821-FHT01B05",
    "GW821-FHT01H02",
    "GW821-FHT02A12",
    "GW821-FHT05F08",
    "GW821-FHT09G11",
    "GW821-FHT10B07",
    "GW821-FHT11D02",
    "GW822-FHT02A07",
    "GW822-FHT02H01",
    "GW822-FHT03E02",
    "GW823-FHT05D11",
}

LUI_NIELSEN_2022_FLYE_UNICYCLER_GENOMES = {
    "FW104-R5.2",
    "FW300-N1B4.2",
    "FW305-C-49.2",
    "FW306-02-B.2",
    "GW822-FHT05A05.2",
    "GW822-FHT05C07.2",
    "GW823-FHT01H08.2",
}

PRICE_2024_HIFI_GENOMES = {
    "EB106-05-01-XG201.3",
    "FW305-C-30-35.4",
    "GW101-3B10.1",
    "GW821-FHT01E09.3",
    "GW821-FHT04C10.3",
    "GW822-FHT03B03.3",
    "GW822-FHT04H01.4",
    "GW823-FHT01F03.4",
}

HARDCODED_GTDB_GENUS_BY_ISOLATE = {
    "FW305-113": "Pseudomonas",
    "GW821-FHT02G11": "Pseudomonas",
}

HARDCODED_SHORT_READ_DATE_BY_GENOME = {
    "GW823-FHT02G05.1": "2024-11-15",
}

HARDCODED_READ_TYPE_BY_READS_ID = {
    "Reads0016183": "single end read",
}

LONG_READ_HARDCODED_METADATA: Dict[str, Dict[str, str]] = {
    "CPT19-411-MTA": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "CPT56D-587-MTF": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "EB106-05-01-XG146": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW104-R8": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW106-PBR-LB-1-21": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW106-PBR-LB-2-19": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW106-PBR-R2A-1-13": {"assembler": "Flye", "illumina": "NovaSeq"},
    "FW106-PBR-R2A-3-15": {"assembler": "Flye", "illumina": "NovaSeq"},
    "FW215-T2": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW300-N1A5": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW300-N1B4": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW300-N2A2": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW300-N2C3": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW300-N2E3": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW300-N2F2": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-113": {"assembler": "Unicycler", "illumina": "HiSeq"},
    "FW305-123": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-127": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-25": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-3": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-3-2-15-A-R2A1": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-3-2-15-E-R2A1": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-3-2-15-F-LB2": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-42": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-47B": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-53": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-63": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-BF8": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-C-10-9": {"assembler": "Flye", "illumina": "N/A"},
    "FW305-C-134A": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-C-21": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-C-271": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW305-C-272": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW306-04-A": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW306-05-C": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW306-06-A": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW306-07-I": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW306-2-11AB": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW306-2-11AD": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW306-2-2C-B10A": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW306-2-2C-D06B": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW507-14TSA": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "FW507-4G11": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW101-11A03": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW101-3H06": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW101-3H11": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW247-26LB": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW456-11-11-14-LB4": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW456-12-1-14-LB2": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW456-12-10-14-TSB1": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW456-E6": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW456-L13": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW456-L15": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW456-R20": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW460-11-11-14-LB1": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW460-11-11-14-LB5": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW460-12-1-14-LB3": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW460-12-10-14-LB2": {"assembler": "Flye", "illumina": "None"},
    "GW460-8": {"assembler": "Flye + Circlator", "illumina": "NovaSeq"},
    "GW460-C8": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW460-E12": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW460-R15": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW531-T4": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW704-E3": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW821-FHT01H03": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW821-FHT02G11": {"assembler": "Flye", "illumina": "HiSeq"},
    "GW821-FHT05B06": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT02A01": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT02H05": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT03B08": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT03C01": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT05A08": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT05C05": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT05D05": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT05E02": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT05E05": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW822-FHT07H11": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW823-FHT01D03": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW823-FHT05C09": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW823-FHT05D11": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "GW823-FHT05G12": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MLSD5-FHT05A06": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MLSD5-FHT05C12": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MPR-LB4": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MPR-R2A7": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MT049": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MT066": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MT094": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MT123": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MT124": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MT58": {"assembler": "Unicycler", "illumina": "NovaSeq"},
    "MT86": {"assembler": "Unicycler", "illumina": "NovaSeq"},
}


def normalize_isolate_key(name: Optional[str]) -> str:
    if not name:
        return ""
    return str(name).split(".", 1)[0].strip()


def _normalized_token(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _is_metadata_unknown(value: Optional[str]) -> bool:
    token = _normalized_token(value)
    return token in {"", "unknown", "na", "none", "null"}


def _assembler_family(value: Optional[str]) -> str:
    text = str(value or "").lower()
    if "unicycler" in text:
        return "unicycler"
    if "flye" in text:
        return "flye"
    if "spades" in text:
        return "spades"
    if "canu" in text:
        return "canu"
    if "metaspades" in text:
        return "metaspades"
    return ""


def _illumina_family(value: Optional[str]) -> str:
    text = str(value or "").lower()
    if "novaseq" in text:
        return "novaseq"
    if "hiseq" in text:
        return "hiseq"
    if "nextseq" in text:
        return "nextseq"
    return ""


def hardcoded_metadata_for_isolate(isolate_name: Optional[str]) -> Dict[str, str]:
    return LONG_READ_HARDCODED_METADATA.get(normalize_isolate_key(isolate_name), {})


def has_hardcoded_long_read_metadata(isolate_name: Optional[str]) -> bool:
    return normalize_isolate_key(isolate_name) in LONG_READ_HARDCODED_METADATA


def normalize_hardcoded_assembler_name(value: Optional[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    family = _assembler_family(text)
    if family == "unicycler":
        return "Unicycler"
    if family == "flye":
        return "Flye"
    if family == "spades":
        return "SPAdes"
    if family == "canu":
        return "CANU"
    if family == "metaspades":
        return "MetaSPAdes"
    return text


def hardcoded_illumina_model(value: Optional[str]) -> str:
    if _is_metadata_unknown(value):
        return ""
    family = _illumina_family(value)
    if family == "novaseq":
        return "Illumina NovaSeq 6000"
    if family == "hiseq":
        return "Illumina HiSeq 4000"
    if family == "nextseq":
        return "Illumina NextSeq 2000"
    return ""


def hardcoded_read_type(
    reads_id: Optional[str], default: Optional[str]
) -> Optional[str]:
    if not reads_id:
        return default
    return HARDCODED_READ_TYPE_BY_READS_ID.get(reads_id, default)


def enforce_assembler_match(
    isolate_name: str,
    berdl_assembler: Optional[str],
    metadata_assembler: Optional[str],
) -> None:
    if _is_metadata_unknown(metadata_assembler):
        return
    berdl_family = _assembler_family(berdl_assembler)
    metadata_family = _assembler_family(metadata_assembler)
    if not berdl_family or not metadata_family:
        return
    if berdl_family != metadata_family:
        raise ValueError(
            "Assembler mismatch for "
            f"{isolate_name}: BERDL={berdl_assembler!r}, hardcoded={metadata_assembler!r}"
        )


def enforce_illumina_match(
    isolate_name: str,
    berdl_platform: Optional[str],
    berdl_instrument_model: Optional[str],
    metadata_illumina: Optional[str],
) -> None:
    metadata_token = _normalized_token(metadata_illumina)
    metadata_family = _illumina_family(metadata_illumina)
    berdl_family = _illumina_family(berdl_instrument_model)
    if not berdl_family:
        return
    if metadata_token in {"none", "na"}:
        log_info(
            "FIXME Illumina mismatch for "
            f"{isolate_name}: BERDL={berdl_instrument_model!r}, hardcoded={metadata_illumina!r}. "
            "Update read-source protocol metadata in BERDL."
        )
        return
    if metadata_family and berdl_family != metadata_family:
        log_info(
            "FIXME Illumina mismatch for "
            f"{isolate_name}: BERDL={berdl_instrument_model!r}, hardcoded={metadata_illumina!r}. "
            "Update read-source protocol metadata in BERDL."
        )


def log_info(message: str) -> None:
    print(f"[info] {message}", file=sys.stderr)


def log_debug(message: str, enabled: bool) -> None:
    if enabled:
        print(f"[debug] {message}", file=sys.stderr)


def normalize_edr_link(link: Optional[str]) -> Optional[str]:
    if not link:
        return None
    normalized = str(link).strip()
    if not normalized:
        return None
    if normalized.startswith(EDR_URL_PREFIX):
        return normalized
    match = re.match(r"^/?enigma-data-repository/(.+)$", normalized, flags=re.IGNORECASE)
    if match:
        rel_path = match.group(1).lstrip("/")
        return f"{EDR_URL_PREFIX.rstrip('/')}/{rel_path}"
    return normalized


def _format_json_for_log(value: Any, fallback: str = "<unavailable>") -> str:
    if value is None:
        return fallback
    try:
        return json.dumps(value, sort_keys=True, default=str)
    except (TypeError, ValueError):
        return str(value)


def log_request_failure(
    exc: requests.HTTPError,
    payload: Optional[Dict[str, Any]] = None,
    path: Optional[str] = None,
) -> None:
    request = getattr(exc, "request", None)
    response = exc.response
    request_url = getattr(request, "url", None) or getattr(response, "url", None)
    if not request_url and path:
        request_url = f"{walk_provenance_module.BASE_URL}{path}"

    log_info(f"HTTP request failed: {exc}")
    log_info(f"Failed URL: {request_url or '<unknown>'}")
    if payload is not None:
        log_info(f"Failed payload: {_format_json_for_log(payload)}")
    if response is not None:
        body = (response.text or "").strip()
        if body:
            if len(body) > 2000:
                body = f"{body[:2000]}... [truncated]"
            log_info(f"Response body: {body}")


def enable_request_failure_logging() -> None:
    if getattr(walk_provenance_module, "_generate_ncbi_post_json_wrapped", False):
        return

    original_post_json = walk_provenance_module.post_json

    def wrapped_post_json(path: str, payload: Dict[str, Any], headers: Dict[str, str]) -> Any:
        try:
            return original_post_json(path, payload, headers)
        except requests.HTTPError as exc:
            log_request_failure(exc, payload=payload, path=path)
            raise

    walk_provenance_module.post_json = wrapped_post_json
    walk_provenance_module._generate_ncbi_post_json_wrapped = True


def get_headers() -> Dict[str, str]:
    token = os.environ.get("KB_AUTH_TOKEN")
    if not token:
        raise RuntimeError("KB_AUTH_TOKEN must be set in the environment.")
    return {"Authorization": f"Bearer {token}"}


def get_table_columns(
    headers: Dict[str, str], table: str, cache: Dict[str, List[str]]
) -> List[str]:
    if table in cache:
        return cache[table]
    columns = get_table_schema(headers, table)
    cache[table] = columns
    return columns


def select_first_row(
    headers: Dict[str, str],
    table: str,
    filters: List[Dict[str, Any]],
    columns: Sequence[str],
) -> Optional[Dict[str, Any]]:
    rows = select_all_rows(headers, table, columns=columns, filters=filters, limit=1)
    return rows[0] if rows else None


def select_row_by_id(
    headers: Dict[str, str],
    table: str,
    obj_id: str,
    columns: Sequence[str],
) -> Optional[Dict[str, Any]]:
    id_col = f"{table}_id"
    filters = [{"column": id_col, "operator": "=", "value": obj_id}]
    return select_first_row(headers, table, filters, columns)


def normalize_protocol_names(protocol_field: Any) -> List[str]:
    if protocol_field is None:
        return []
    if isinstance(protocol_field, list):
        return [str(p).strip() for p in protocol_field if str(p).strip()]
    if isinstance(protocol_field, str):
        parts = [p.strip() for p in protocol_field.split(",")]
        return [p for p in parts if p]
    return [str(protocol_field).strip()]


def parse_protocol_metadata(protocol_name: str, description: str) -> Dict[str, Optional[str]]:
    name_lower = (protocol_name or "").lower()
    desc_lower = (description or "").lower()

    sequencing_tech = None
    machine_type = None
    program_version = None

    if "novaseq" in name_lower or "novaseq" in desc_lower:
        if "xplus" in name_lower or "x plus" in desc_lower:
            machine_type = "NovaSeq X Plus"
        elif "6000" in name_lower or "6000" in desc_lower:
            machine_type = "NovaSeq 6000"
        else:
            machine_type = "NovaSeq"
        sequencing_tech = "Illumina"
    elif "hiseq" in name_lower or "hiseq" in desc_lower:
        machine_type = "HiSeq 4000"
        sequencing_tech = "Illumina"
    elif "nextseq" in name_lower or "nextseq" in desc_lower:
        if "2000" in name_lower or "2000" in desc_lower:
            machine_type = "NextSeq 2000"
        elif "500" in name_lower or "500" in desc_lower:
            machine_type = "NextSeq 500"
        else:
            machine_type = "NextSeq"
        sequencing_tech = "Illumina"
    elif "promethion" in name_lower or "promethion" in desc_lower:
        machine_type = "PromethION"
        sequencing_tech = "Oxford Nanopore"
    elif "pacbio" in name_lower or "pacbio" in desc_lower:
        sequencing_tech = "PacBio"
        if "hifi" in name_lower or "hifi" in desc_lower:
            machine_type = "PacBio HiFi"
    elif "minion" in name_lower or "minion" in desc_lower:
        machine_type = "MinION"
        sequencing_tech = "Oxford Nanopore"

    version_patterns = [
        (r"spades\s+v?(\d+\.\d+\.\d+)", "spades"),
        (r"cutadapt\s+v?(\d+\.\d+)", "cutadapt"),
        (r"trimmomatic\s+v?(\d+\.\d+)", "trimmomatic"),
        (r"prokka\s+v?(\d+\.\d+)", "prokka"),
        (r"prodigal", "prodigal"),
        (r"flye\s+v?(\d+\.\d+)", "flye"),
        (r"canu\s+v?(\d+\.\d+)", "canu"),
        (r"metaspades", "metaspades"),
        (r"unicycler", "unicycler"),
    ]

    for pattern, program in version_patterns:
        match = re.search(pattern, description or "", re.IGNORECASE)
        if match:
            if program == "prodigal":
                program_version = "Prodigal"
            elif program == "metaspades":
                program_version = "MetaSPAdes"
            elif program == "unicycler":
                program_version = "Unicycler"
            else:
                version = match.group(1) if match.groups() else ""
                program_version = f"{program.capitalize()} {version}".strip()
            break

    return {
        "sequencing_technology": sequencing_tech,
        "machine_type": machine_type,
        "program_version": program_version,
    }


def is_paired_read_type(read_type: Optional[str]) -> bool:
    if not read_type:
        return False
    read_type_lower = read_type.lower()
    return "paired" in read_type_lower or "pair" in read_type_lower


def get_protocol_info(
    headers: Dict[str, str],
    protocol_name: str,
    column_cache: Dict[str, List[str]],
    info_cache: Dict[str, Dict[str, Optional[str]]],
) -> Dict[str, Optional[str]]:
    if protocol_name in info_cache:
        return info_cache[protocol_name]

    columns = get_table_columns(headers, PROTOCOL_TABLE, column_cache)
    desired = [col for col in ["sdt_protocol_name", "sdt_protocol_description", "link"] if col in columns]
    if not desired:
        info_cache[protocol_name] = {}
        return info_cache[protocol_name]

    row = select_first_row(
        headers,
        PROTOCOL_TABLE,
        [{"column": "sdt_protocol_name", "operator": "=", "value": protocol_name}],
        desired,
    )
    if row is None and "sdt_protocol_id" in columns:
        row = select_first_row(
            headers,
            PROTOCOL_TABLE,
            [{"column": "sdt_protocol_id", "operator": "=", "value": protocol_name}],
            desired,
        )

    name = row.get("sdt_protocol_name") if row else protocol_name
    description = row.get("sdt_protocol_description") if row else ""
    metadata = parse_protocol_metadata(name or protocol_name, description or "")
    metadata["protocol_name"] = name or protocol_name
    metadata["protocol_description"] = description or ""
    metadata["protocol_link"] = row.get("link") if row else None
    info_cache[protocol_name] = metadata
    return metadata


def find_oldest_reads_with_fastq(
    genome_token: str,
    out_lookup: Dict[str, List[Dict[str, Any]]],
    downstream_lookup: Dict[str, List[Dict[str, Any]]],
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    read_cache: Dict[str, Dict[str, Any]],
    strain_token: Optional[str] = None,
    log_label: Optional[str] = None,
    genome_name: Optional[str] = None,
) -> List[Dict[str, Any]]:
    downstream_genome_cache: Dict[str, set[str]] = {}
    upstream_copy_reads_cache: Dict[str, set[str]] = {}

    def read_link_ok(link: Optional[str]) -> bool:
        normalized_link = normalize_edr_link(link)
        if not normalized_link:
            return False
        link_lower = normalized_link.lower()
        if FASTQ_HOST not in link_lower:
            return False
        return ".fastq" in link_lower or ".fq" in link_lower

    def get_reads_data(obj_id: str) -> Dict[str, Any]:
        if obj_id in read_cache:
            return read_cache[obj_id]
        columns = get_table_columns(headers, "sdt_reads", column_cache)
        desired = [
            col
            for col in [
                "sdt_reads_id",
                "sdt_reads_name",
                "link",
                "read_type_sys_oterm_name",
                "sequencing_technology_sys_oterm_name",
            ]
            if col in columns
        ]
        row = select_row_by_id(headers, "sdt_reads", obj_id, desired)
        reads_data = row or {}
        read_cache[obj_id] = reads_data
        return reads_data

    def reads_process_flags(reads_token: str) -> Tuple[bool, bool, bool]:
        produced_by_copy = False
        produced_by_reads_processing = False
        processed = False
        for proc in out_lookup.get(reads_token, []):
            process_name = (proc.get("process_term_name") or "").lower()
            if "copy data" in process_name:
                produced_by_copy = True
            if "reads processing" in process_name:
                produced_by_reads_processing = True
            if (
                "reads processing" in process_name
                or "cutadapt" in process_name
                or "trimmomatic" in process_name
            ):
                processed = True
        return produced_by_copy, produced_by_reads_processing, processed

    def has_processed_name_marker(value: Optional[str]) -> bool:
        if not value:
            return False
        value_lower = value.lower()
        markers = [
            "cutadapt",
            "trimmomatic",
            "-cut",
            "_cut",
            ".cut",
            "-trim",
            "_trim",
            ".trim",
            "reads_cut",
            "reads_trim",
        ]
        return any(marker in value_lower for marker in markers)

    def collect_reads_protocols(reads_token: str) -> List[str]:
        shotgun_protocols: List[str] = []
        other_protocols: List[str] = []

        def add_protocols(proc: Dict[str, Any]) -> None:
            process_name = (proc.get("process_term_name") or "").lower()
            normalized = normalize_protocol_names(proc.get("protocol"))
            if not normalized:
                return
            if (
                "shotgun sequencing and assembly" in process_name
                or "shotgun sequencing" in process_name
                or "sequencing" in process_name
            ):
                shotgun_protocols.extend(normalized)
            else:
                other_protocols.extend(normalized)

        # Include protocols from the selected reads object and its upstream reads lineage
        # so copied reads can still inherit the original sequencing protocol.
        visited_reads: set[str] = set()

        def walk_upstream_reads(token: str) -> None:
            if token in visited_reads:
                return
            visited_reads.add(token)
            for proc in out_lookup.get(token, []):
                add_protocols(proc)
                for inp in proc.get("input_objs", []):
                    inp_table, _ = parse_token(inp)
                    if inp_table == "sdt_reads":
                        walk_upstream_reads(inp)

        walk_upstream_reads(reads_token)

        for proc in downstream_lookup.get(reads_token, []):
            add_protocols(proc)

        ordered = shotgun_protocols + other_protocols
        seen: set[str] = set()
        unique = []
        for name in ordered:
            if name in seen:
                continue
            seen.add(name)
            unique.append(name)
        return unique

    def collect_reads_inputs(start_token: str) -> List[str]:
        visited: set[str] = set()
        reads_inputs: List[str] = []

        def walk_upstream(obj_token: str) -> None:
            if obj_token in visited:
                return
            visited.add(obj_token)

            table_name, obj_id = parse_token(obj_token)
            if not table_name or not obj_id:
                return

            if table_name == "sdt_reads":
                reads_inputs.append(obj_token)
                return

            proc_list = out_lookup.get(obj_token, [])
            for proc in proc_list:
                process_name = (proc.get("process_term_name") or "").lower()
                if table_name in ["sdt_assembly", "sdt_genome"]:
                    for inp in proc.get("input_objs", []):
                        walk_upstream(inp)
                else:
                    for inp in proc.get("input_objs", []):
                        inp_table, _ = parse_token(inp)
                        if inp_table in ["sdt_reads", "sdt_assembly"]:
                            walk_upstream(inp)

        walk_upstream(start_token)
        seen: set[str] = set()
        unique_reads = []
        for token in reads_inputs:
            if token in seen:
                continue
            seen.add(token)
            unique_reads.append(token)
        return unique_reads

    def collect_ancestral_reads(start_reads_token: str) -> List[str]:
        visited: set[str] = set()
        ancestors: List[str] = []

        def walk_upstream(reads_token: str) -> None:
            if reads_token in visited:
                return
            visited.add(reads_token)

            table_name, _ = parse_token(reads_token)
            if table_name != "sdt_reads":
                return

            upstream_reads: List[str] = []
            for proc in out_lookup.get(reads_token, []):
                process_name = (proc.get("process_term_name") or "").lower()
                if "reads processing" in process_name or "copy data" in process_name:
                    for inp in proc.get("input_objs", []):
                        inp_table, _ = parse_token(inp)
                        if inp_table == "sdt_reads":
                            upstream_reads.append(inp)

            if not upstream_reads:
                ancestors.append(reads_token)
                return

            for inp in upstream_reads:
                walk_upstream(inp)

        walk_upstream(start_reads_token)
        seen: set[str] = set()
        unique_reads = []
        for token in ancestors:
            if token in seen:
                continue
            seen.add(token)
            unique_reads.append(token)
        return unique_reads

    def collect_reads_downstream(
        start_token: str, only_copy: bool
    ) -> List[Dict[str, Any]]:
        visited: set[str] = set()
        reads_candidates: List[Dict[str, Any]] = []

        def walk_downstream(
            obj_token: str,
            depth: int = 0,
            path: Optional[List[str]] = None,
            parent_token: Optional[str] = None,
            parent_process: Optional[str] = None,
            parent_process_id: Optional[str] = None,
        ) -> None:
            if path is None:
                path = []
            if obj_token in visited:
                return
            visited.add(obj_token)

            table_name, obj_id = parse_token(obj_token)
            if not table_name or not obj_id:
                return

            current_path = path + [obj_token]

            if table_name == "sdt_reads":
                reads_data = get_reads_data(obj_id)
                link = normalize_edr_link(reads_data.get("link"))
                if read_link_ok(link):
                    protocols = collect_reads_protocols(obj_token)
                    (
                        produced_by_copy,
                        produced_by_reads_processing,
                        processed,
                    ) = reads_process_flags(obj_token)
                    reads_name = reads_data.get("sdt_reads_name")
                    processed_by_name = has_processed_name_marker(reads_name) or has_processed_name_marker(
                        link
                    )
                    reads_candidates.append(
                        {
                            "reads_id": obj_id,
                            "reads_name": reads_name,
                            "link": link,
                            "read_type": hardcoded_read_type(
                                obj_id, reads_data.get("read_type_sys_oterm_name")
                            ),
                            "sequencing_technology": reads_data.get(
                                "sequencing_technology_sys_oterm_name"
                            ),
                            "protocols": sorted(set(protocols)),
                            "depth": depth,
                            "path": current_path.copy(),
                            "produced_by_copy_data": produced_by_copy,
                            "produced_by_reads_processing": produced_by_reads_processing,
                            "processed_by_trim": processed,
                            "processed_name_marker": processed_by_name,
                            "reads_token": obj_token,
                            "parent_reads_token": parent_token,
                            "parent_process_name": parent_process,
                            "parent_process_id": parent_process_id,
                        }
                    )

            for proc in downstream_lookup.get(obj_token, []):
                process_name = (proc.get("process_term_name") or "").lower()
                if only_copy and "copy data" not in process_name:
                    continue
                out_token = proc.get("output_obj")
                if out_token:
                    walk_downstream(
                        out_token,
                        depth + 1,
                        current_path,
                        parent_token=obj_token,
                        parent_process=proc.get("process_term_name"),
                        parent_process_id=proc.get("sys_process_id"),
                    )

        walk_downstream(start_token)
        return reads_candidates

    def log_reads(message: str, reads_list: List[Dict[str, Any]]) -> None:
        prefix = log_label or ""
        if not reads_list:
            log_info(f"{prefix}{message}: none")
            return
        details = ", ".join(
            f"{r.get('reads_id')}:{r.get('reads_name')}" for r in reads_list
        )
        log_info(f"{prefix}{message}: {details}")

    def log_reads_detail(message: str, reads_list: List[Dict[str, Any]]) -> None:
        prefix = log_label or ""
        log_info(f"{prefix}{message}: {len(reads_list)}")
        for reads in reads_list:
            log_info(
                f"{prefix}  - id={reads.get('reads_id')} "
                f"name={reads.get('reads_name')!r} "
                f"read_type={reads.get('read_type')!r} "
                f"produced_by_copy={reads.get('produced_by_copy_data')} "
                f"produced_by_reads_processing={reads.get('produced_by_reads_processing')} "
                f"processed_by_trim={reads.get('processed_by_trim')} "
                f"processed_name_marker={reads.get('processed_name_marker')} "
                f"parent_reads={reads.get('parent_reads_token')!r} "
                f"parent_process={reads.get('parent_process_name')!r} "
                f"parent_process_id={reads.get('parent_process_id')!r} "
                f"link={reads.get('link')!r}"
            )

    def choose_paired_from_group(reads_group: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if len(reads_group) <= 1:
            return reads_group
        r1 = None
        r2 = None
        for reads in reads_group:
            name_lower = (reads.get("reads_name") or "").lower()
            link_lower = (reads.get("link") or "").lower()
            marker = f"{name_lower} {link_lower}"
            if any(x in marker for x in ["_r1", "_1.fastq", "_1.fq"]):
                r1 = reads
            elif any(x in marker for x in ["_r2", "_2.fastq", "_2.fq"]):
                r2 = reads
        if r1 and r2:
            return [r1, r2]
        return reads_group

    def get_reads_protocols(reads_token: str) -> List[str]:
        return collect_reads_protocols(reads_token)

    def reads_is_long(reads: Dict[str, Any]) -> bool:
        seq_tech = reads.get("sequencing_technology") or ""
        return is_long_read_tech(None, str(seq_tech))

    def reads_is_short(reads: Dict[str, Any]) -> bool:
        return not reads_is_long(reads)

    def reads_is_nanopore(reads: Dict[str, Any]) -> bool:
        seq_tech = str(reads.get("sequencing_technology") or "").lower()
        marker = " ".join(
            [
                str(reads.get("reads_name") or ""),
                str(reads.get("link") or ""),
            ]
        ).lower()
        return "nanopore" in seq_tech or "/ont/" in marker or "ont/" in marker

    def reads_is_pacbio(reads: Dict[str, Any]) -> bool:
        seq_tech = str(reads.get("sequencing_technology") or "").lower()
        marker = " ".join(
            [
                str(reads.get("reads_name") or ""),
                str(reads.get("link") or ""),
            ]
        ).lower()
        return "pacbio" in seq_tech or "/pacbio/" in marker or "pacbio/" in marker

    def prefer_nanopore_when_long_platforms_mixed(
        reads_list: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        if not reads_list:
            return reads_list
        has_nanopore = any(reads_is_long(r) and reads_is_nanopore(r) for r in reads_list)
        has_pacbio = any(reads_is_long(r) and reads_is_pacbio(r) for r in reads_list)
        if not (has_nanopore and has_pacbio):
            return reads_list
        filtered = [
            r
            for r in reads_list
            if not (reads_is_long(r) and reads_is_pacbio(r))
        ]
        log_info(
            f"{log_label or ''}WARNING: both PacBio and Nanopore long reads were selected "
            f"for {genome_token}; keeping Nanopore long reads and dropping PacBio long reads."
        )
        log_reads_detail("Reads after preferring Nanopore over PacBio", filtered)
        return filtered

    def reads_has_raw_paired_marker(reads: Dict[str, Any]) -> bool:
        marker = " ".join(
            [
                str(reads.get("reads_name") or ""),
                str(reads.get("link") or ""),
            ]
        ).lower()
        return bool(re.search(r"raw[_-]?paired", marker))

    def reads_inferred_date(reads: Dict[str, Any]) -> Optional[str]:
        return extract_date_from_filenames(
            Path(str(reads.get("link") or "")).name,
            str(reads.get("reads_name") or ""),
        )

    def find_downstream_genome_tokens(start_token: Optional[str]) -> set[str]:
        if not start_token:
            return set()
        cached = downstream_genome_cache.get(start_token)
        if cached is not None:
            return cached
        found: set[str] = set()
        visited: set[str] = set()

        def walk(token: str) -> None:
            if token in visited:
                return
            visited.add(token)
            table_name, _ = parse_token(token)
            if table_name == "sdt_genome":
                found.add(token)
            for proc in downstream_lookup.get(token, []):
                output_obj = proc.get("output_obj")
                if output_obj:
                    walk(str(output_obj))

        walk(start_token)
        downstream_genome_cache[start_token] = found
        return found

    def find_upstream_copy_reads_tokens(start_token: Optional[str]) -> set[str]:
        if not start_token:
            return set()
        cached = upstream_copy_reads_cache.get(start_token)
        if cached is not None:
            return cached
        found: set[str] = set()
        visited: set[str] = set()

        def walk(token: str) -> None:
            if token in visited:
                return
            visited.add(token)
            table_name, _ = parse_token(token)
            if table_name != "sdt_reads":
                return
            for proc in out_lookup.get(token, []):
                process_name = (proc.get("process_term_name") or "").lower()
                if "copy data" not in process_name:
                    continue
                for inp in proc.get("input_objs", []) or []:
                    inp_table, _ = parse_token(inp)
                    if inp_table != "sdt_reads":
                        continue
                    found.add(inp)
                    walk(inp)

        walk(start_token)
        upstream_copy_reads_cache[start_token] = found
        return found

    def select_best_reads(candidates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if not candidates:
            return []
        if len(candidates) == 1:
            return candidates
        grouped: Dict[Tuple[Optional[str], str, Optional[str]], List[Dict[str, Any]]] = defaultdict(list)
        for reads in candidates:
            parent_token = reads.get("parent_reads_token")
            parent_process = (reads.get("parent_process_name") or "").lower()
            parent_process_id = reads.get("parent_process_id")
            grouped[(parent_token, parent_process, parent_process_id)].append(reads)
        prioritized_groups = {
            key: group
            for key, group in grouped.items()
            if "copy data" in key[1] or "shotgun sequencing" in key[1]
        }
        if prioritized_groups:
            grouped = prioritized_groups

        has_long_candidates = any(reads_is_long(reads) for reads in candidates)
        short_paired_groups: List[
            Tuple[Tuple[Optional[str], str, Optional[str]], List[Dict[str, Any]]]
        ] = []

        def select_long_only_candidates() -> List[Dict[str, Any]]:
            long_only = [reads for reads in candidates if reads_is_long(reads)]
            if long_only:
                log_reads_detail(
                    "Selected long-read candidates only; no short-read candidates remained",
                    long_only,
                )
            return long_only

        def is_short_read_candidate_group(reads_group: List[Dict[str, Any]]) -> bool:
            if not reads_group:
                return False
            if not all(reads_is_short(reads) for reads in reads_group):
                return False
            if len(reads_group) >= 2:
                return True
            return any(is_paired_read_type(reads.get("read_type")) for reads in reads_group)

        for key, group in grouped.items():
            pair = choose_paired_from_group(group)
            # Treat mixed groups as valid short-read candidates by evaluating
            # the short-read subset independently (for example, clean_paired + ONT).
            short_pair = [reads for reads in pair if reads_is_short(reads)]
            if is_short_read_candidate_group(short_pair):
                short_paired_groups.append((key, short_pair))

        forced_short_read_date = HARDCODED_SHORT_READ_DATE_BY_GENOME.get(genome_name or "")
        if forced_short_read_date:
            matched_forced_groups = []
            for _, pair in short_paired_groups:
                short_dates = {reads_inferred_date(reads) for reads in pair}
                short_dates.discard(None)
                if forced_short_read_date in short_dates:
                    matched_forced_groups.append(pair)
            if len(matched_forced_groups) == 1:
                selected_pair = matched_forced_groups[0]
                log_reads_detail(
                    f"Selected hardcoded short-read date {forced_short_read_date}",
                    selected_pair,
                )
                return selected_pair
            if len(matched_forced_groups) > 1:
                log_info(
                    f"{log_label or ''}WARNING: hardcoded short-read date "
                    f"{forced_short_read_date} matched multiple candidate groups for "
                    f"{genome_token}."
                )
            else:
                log_info(
                    f"{log_label or ''}WARNING: hardcoded short-read date "
                    f"{forced_short_read_date} had no matching short-read pair for "
                    f"{genome_token}."
                )

        ambiguous_short_candidates = len(short_paired_groups) > 1
        if ambiguous_short_candidates:
            # Rule 1: in hybrid contexts, prefer raw_paired short-read pairs
            # that match a long-read date, if present.
            long_read_dates = {
                reads_inferred_date(reads)
                for reads in candidates
                if reads_is_long(reads)
            }
            long_read_dates.discard(None)
            raw_paired_short_groups = [
                (key, pair)
                for key, pair in short_paired_groups
                if any(reads_has_raw_paired_marker(reads) for reads in pair)
            ]
            date_matched_raw_paired_groups = []
            if long_read_dates:
                for key, pair in raw_paired_short_groups:
                    short_dates = {reads_inferred_date(reads) for reads in pair}
                    short_dates.discard(None)
                    if short_dates & long_read_dates:
                        date_matched_raw_paired_groups.append((key, pair))
            if has_long_candidates and len(date_matched_raw_paired_groups) == 1:
                selected_pair = date_matched_raw_paired_groups[0][1]
                log_reads_detail(
                    "Selected date-matched raw_paired short-read pair for hybrid read context",
                    selected_pair,
                )
                return selected_pair

            # Rule 2: remove short-read candidates whose provenance reaches non-target genomes.
            filtered_short_groups: List[
                Tuple[Tuple[Optional[str], str, Optional[str]], List[Dict[str, Any]]]
            ] = []
            eliminated_groups = 0
            for key, pair in short_paired_groups:
                downstream_genomes: set[str] = set()
                for reads in pair:
                    reads_token = str(reads.get("reads_token") or "")
                    related_tokens: set[str] = {reads_token}
                    related_tokens.update(find_upstream_copy_reads_tokens(reads_token))
                    for token in related_tokens:
                        downstream_genomes.update(find_downstream_genome_tokens(token))
                other_genomes = {
                    genome_ref for genome_ref in downstream_genomes if genome_ref != genome_token
                }
                if other_genomes:
                    eliminated_groups += 1
                    continue
                filtered_short_groups.append((key, pair))

            if eliminated_groups:
                log_info(
                    f"{log_label or ''}Filtered {eliminated_groups} short-read candidate "
                    "group(s) due to downstream provenance linking to other genomes."
                )
            if not filtered_short_groups:
                log_info(
                    f"{log_label or ''}WARNING: short-read ambiguity for {genome_token}; "
                    "downstream-genome filtering removed all short-read candidates."
                )
                # Do not fall back to arbitrary short-read selection when all
                # ambiguous short-read groups were eliminated.
                return select_long_only_candidates()
            elif len(filtered_short_groups) == 1:
                selected_pair = filtered_short_groups[0][1]
                log_reads_detail(
                    "Selected short-read pair after downstream-genome filtering",
                    selected_pair,
                )
                return selected_pair
            else:
                log_info(
                    f"{log_label or ''}WARNING: short-read ambiguity remains for {genome_token} "
                    "after raw_paired date-match preference and downstream-genome filtering."
                )

        best_group = None
        best_pair = []
        for parent_token, group in grouped.items():
            pair = choose_paired_from_group(group)
            if len(pair) >= 2:
                best_group = group
                best_pair = pair
                break
        if best_pair:
            log_reads_detail("Selected paired reads from same parent", best_pair)
            return best_pair
        if len(grouped) == 1:
            only_group = next(iter(grouped.values()))
            return only_group
        best_group = max(grouped.values(), key=len)
        log_reads_detail("Selected reads from largest parent group", best_group)
        return best_group

    def is_filtlong_long_read(reads: Dict[str, Any]) -> bool:
        if not reads_is_long(reads):
            return False
        marker = " ".join(
            [
                str(reads.get("reads_name") or ""),
                str(reads.get("link") or ""),
            ]
        ).lower()
        return "filtlong" in marker

    def drop_filtlong_when_raw_long_available(
        candidates: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        if not candidates:
            return candidates
        has_raw_long = any(
            reads_is_long(r) and not is_filtlong_long_read(r) for r in candidates
        )
        if not has_raw_long:
            return candidates
        filtered = [
            r
            for r in candidates
            if not (reads_is_long(r) and is_filtlong_long_read(r))
        ]
        if len(filtered) != len(candidates):
            log_reads_detail(
                "Dropped filtlong long reads because raw long reads are available",
                filtered,
            )
        return filtered

    def merge_with_long_read_candidates(
        selected: List[Dict[str, Any]], candidates: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        if not candidates:
            return selected
        selected_has_long = any(reads_is_long(r) for r in selected)
        if selected_has_long:
            return selected
        long_candidates = [r for r in candidates if reads_is_long(r)]
        if not long_candidates:
            return selected
        merged = list(selected)
        seen_ids = {r.get("reads_id") for r in merged if r.get("reads_id")}
        for reads in long_candidates:
            reads_id = reads.get("reads_id")
            if reads_id and reads_id in seen_ids:
                continue
            merged.append(reads)
            if reads_id:
                seen_ids.add(reads_id)
        log_reads_detail("Added long-read candidates to selected reads", merged)
        return merged

    def select_submission_reads(
        ancestor_reads_token: str,
        ancestor_protocols: List[str],
        ancestor_seq_tech: Optional[str],
    ) -> List[Dict[str, Any]]:
        def is_raw_candidate(reads: Dict[str, Any]) -> bool:
            if reads.get("produced_by_reads_processing"):
                return False
            if reads.get("processed_by_trim"):
                return False
            if reads.get("processed_name_marker"):
                return False
            return True

        # Step 2: follow only Copy Data processes downstream from ancestral reads.
        copy_candidates = collect_reads_downstream(ancestor_reads_token, only_copy=True)
        if copy_candidates:
            copy_candidates = drop_filtlong_when_raw_long_available(copy_candidates)
            log_reads_detail("Copy-data FASTQ candidates", copy_candidates)
            chosen = select_best_reads(copy_candidates)
        else:
            # Step 3 fallback: any raw FASTQ reads in EDR, excluding processed reads.
            all_candidates = collect_reads_downstream(ancestor_reads_token, only_copy=False)
            raw_candidates = [reads for reads in all_candidates if is_raw_candidate(reads)]
            raw_candidates = drop_filtlong_when_raw_long_available(raw_candidates)
            log_reads_detail("Raw downstream FASTQ candidates", raw_candidates)
            chosen = select_best_reads(raw_candidates)

        for reads in chosen:
            reads["source_reads_token"] = ancestor_reads_token
            reads["source_protocols"] = ancestor_protocols
            reads["source_sequencing_technology"] = ancestor_seq_tech
        return chosen

    reads_inputs = collect_reads_inputs(genome_token)
    if reads_inputs:
        log_info(f"{log_label or ''}Assembly reads inputs: {len(reads_inputs)}")
    selected_reads: List[Dict[str, Any]] = []
    for reads_input in reads_inputs:
        ancestral_reads = collect_ancestral_reads(reads_input)
        if not ancestral_reads:
            continue
        for ancestor_token in ancestral_reads:
            table_name, obj_id = parse_token(ancestor_token)
            if table_name != "sdt_reads" or not obj_id:
                continue
            ancestor_data = get_reads_data(obj_id)
            ancestor_protocols = get_reads_protocols(ancestor_token)
            ancestor_seq_tech = ancestor_data.get("sequencing_technology_sys_oterm_name")
            submissions = select_submission_reads(
                ancestor_token, ancestor_protocols, ancestor_seq_tech
            )
            if submissions:
                selected_reads.extend(submissions)

    if selected_reads:
        seen_ids: set[str] = set()
        unique_reads: List[Dict[str, Any]] = []
        for reads in selected_reads:
            reads_id = reads.get("reads_id")
            if reads_id in seen_ids:
                continue
            if reads_id:
                seen_ids.add(reads_id)
            unique_reads.append(reads)
        unique_reads = prefer_nanopore_when_long_platforms_mixed(unique_reads)
        log_reads("Selected reads", unique_reads)
        return unique_reads

    if strain_token:
        downstream_candidates = collect_reads_downstream(strain_token, only_copy=False)
        raw_downstream = [
            reads
            for reads in downstream_candidates
            if not reads.get("produced_by_reads_processing")
            and not reads.get("processed_by_trim")
            and not reads.get("processed_name_marker")
        ]
        # If assembly provenance indicates short-read-only sequencing, do not
        # pull long reads from the broad strain-downstream fallback path.
        assembly_processes = find_assembly_processes(genome_token, out_lookup)
        assembly_protocols = collect_protocols_from_processes(assembly_processes)
        assembly_protocol_cache: Dict[str, Dict[str, Optional[str]]] = {}
        assembly_seq_techs = infer_sequencing_techs(
            assembly_protocols, headers, column_cache, assembly_protocol_cache
        )
        if assembly_seq_techs and not any(
            is_long_read_tech(None, tech) for tech in assembly_seq_techs
        ):
            raw_downstream = [reads for reads in raw_downstream if not reads_is_long(reads)]
        if raw_downstream:
            raw_downstream = drop_filtlong_when_raw_long_available(raw_downstream)
            log_reads("Found raw strain-downstream reads", raw_downstream)
            chosen = select_best_reads(raw_downstream)
            chosen = merge_with_long_read_candidates(chosen, raw_downstream)
            return prefer_nanopore_when_long_platforms_mixed(chosen)

    return []


def find_samples_from_genome(
    genome_token: str,
    out_lookup: Dict[str, List[Dict[str, Any]]],
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
) -> List[Dict[str, Any]]:
    visited: set[str] = set()
    samples_found: List[Dict[str, Any]] = []

    def walk_upstream(obj_token: str, current_protocol: Optional[str] = None) -> None:
        if obj_token in visited:
            return
        visited.add(obj_token)

        table_name, obj_id = parse_token(obj_token)
        if not table_name or not obj_id:
            return

        if table_name == "sdt_sample":
            columns = get_table_columns(headers, "sdt_sample", column_cache)
            desired = [
                col
                for col in [
                    "sdt_sample_id",
                    "sdt_sample_name",
                    "sdt_location_name",
                    "date",
                    "depth_meter",
                    "material_sys_oterm_name",
                    "sdt_sample_description",
                ]
                if col in columns
            ]
            sample_row = select_row_by_id(headers, "sdt_sample", obj_id, desired)
            if sample_row:
                samples_found.append(
                    {
                        "sample_id": obj_id,
                        "sample_name": sample_row.get("sdt_sample_name"),
                        "sample_token": f"sdt_sample:{obj_id}",
                        "location_name": sample_row.get("sdt_location_name"),
                        "date": sample_row.get("date"),
                        "depth_meter": sample_row.get("depth_meter"),
                        "material_name": sample_row.get("material_sys_oterm_name"),
                        "description": sample_row.get("sdt_sample_description"),
                        "protocol": current_protocol,
                    }
                )

        for proc in out_lookup.get(obj_token, []):
            protocol = proc.get("protocol") or current_protocol
            for inp in proc.get("input_objs", []):
                if table_name == "sdt_sample":
                    walk_upstream(inp, protocol)
                else:
                    walk_upstream(inp, current_protocol)

    walk_upstream(genome_token)

    seen_ids: set[str] = set()
    unique_samples = []
    for sample in samples_found:
        if sample["sample_id"] in seen_ids:
            continue
        seen_ids.add(sample["sample_id"])
        unique_samples.append(sample)
    return unique_samples


def get_location_info(
    headers: Dict[str, str],
    location_name: str,
    column_cache: Dict[str, List[str]],
) -> Optional[Dict[str, Any]]:
    if not location_name:
        return None
    columns = get_table_columns(headers, "sdt_location", column_cache)
    desired = [
        col
        for col in [
            "sdt_location_name",
            "latitude_degree",
            "longitude_degree",
            "country_sys_oterm_name",
            "region",
            "biome_sys_oterm_name",
        ]
        if col in columns
    ]
    row = select_first_row(
        headers,
        "sdt_location",
        [{"column": "sdt_location_name", "operator": "=", "value": location_name}],
        desired,
    )
    if not row:
        return None
    return {
        "location_name": location_name,
        "latitude": row.get("latitude_degree"),
        "longitude": row.get("longitude_degree"),
        "country": row.get("country_sys_oterm_name"),
        "region": row.get("region"),
        "biome": row.get("biome_sys_oterm_name"),
    }


def get_strain_info(
    headers: Dict[str, str],
    strain_name: str,
    column_cache: Dict[str, List[str]],
) -> Optional[Dict[str, Any]]:
    if not strain_name:
        return None
    columns = get_table_columns(headers, "sdt_strain", column_cache)
    desired = [
        col for col in ["sdt_strain_name", "sdt_strain_description"] if col in columns
    ]
    row = select_first_row(
        headers,
        "sdt_strain",
        [{"column": "sdt_strain_name", "operator": "=", "value": strain_name}],
        desired,
    )
    if not row:
        return None
    return {
        "strain_name": row.get("sdt_strain_name") or strain_name,
        "description": row.get("sdt_strain_description"),
    }


def get_strain_id(
    headers: Dict[str, str],
    strain_name: str,
    column_cache: Dict[str, List[str]],
) -> Optional[str]:
    if not strain_name:
        return None
    columns = get_table_columns(headers, "sdt_strain", column_cache)
    if "sdt_strain_id" not in columns or "sdt_strain_name" not in columns:
        return None
    row = select_first_row(
        headers,
        "sdt_strain",
        [{"column": "sdt_strain_name", "operator": "=", "value": strain_name}],
        ["sdt_strain_id"],
    )
    return row.get("sdt_strain_id") if row else None


def load_biosample_template_workbook(output_dir: str) -> Tuple[Path, Any, int, Dict[str, int]]:
    candidates = list(BIOSAMPLE_TEMPLATE_CANDIDATES)
    candidates.extend(
        [Path(output_dir) / "Microbe.1.0.xlsx", Path(output_dir) / "MIcrobe.1.0.xlsx"]
    )
    template_path = next((path for path in candidates if path.exists()), None)
    if not template_path:
        raise FileNotFoundError(
            "BioSample template not found. Expected Microbe.1.0.xlsx or MIcrobe.1.0.xlsx."
        )
    workbook = load_workbook(template_path)
    sheet = workbook.active

    header_row = None
    header_map: Dict[str, int] = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if (cell.value or "").strip() == "*sample_name":
                header_row = cell.row
                for header_cell in sheet[header_row]:
                    header = header_cell.value
                    if header is None:
                        continue
                    header_str = str(header).strip()
                    if header_str:
                        header_map[header_str] = header_cell.column
                break
        if header_row is not None:
            break

    if header_row is None:
        raise ValueError(f"No header row found in BioSample template: {template_path}")

    return template_path, sheet, header_row, header_map


def load_sra_template_workbook(output_dir: str) -> Tuple[Path, Any, int, Dict[str, int]]:
    candidates = list(SRA_TEMPLATE_CANDIDATES)
    candidates.append(Path(output_dir) / "SRA_metadata.xlsx")
    template_path = next((path for path in candidates if path.exists()), None)
    if not template_path:
        raise FileNotFoundError(
            "SRA template not found. Expected SRA_metadata.xlsx."
        )
    workbook = load_workbook(template_path)
    sheet = workbook["SRA_data"] if "SRA_data" in workbook.sheetnames else workbook.active

    header_row = None
    header_map: Dict[str, int] = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if (cell.value or "").strip() == "sample_name":
                header_row = cell.row
                for header_cell in sheet[header_row]:
                    header = header_cell.value
                    if header is None:
                        continue
                    header_str = str(header).strip()
                    if header_str:
                        header_map[header_str] = header_cell.column
                break
        if header_row is not None:
            break

    if header_row is None:
        raise ValueError(f"No header row found in SRA template: {template_path}")

    return template_path, sheet, header_row, header_map


def load_genome_template_workbook(output_dir: str) -> Tuple[Path, Any, int, Dict[str, int]]:
    candidates = list(GENOME_TEMPLATE_CANDIDATES)
    candidates.append(Path(output_dir) / "Template_GenomeBatch.xlsx")
    template_path = next((path for path in candidates if path.exists()), None)
    if not template_path:
        raise FileNotFoundError(
            "Genome template not found. Expected Template_GenomeBatch.xlsx."
        )
    workbook = load_workbook(template_path)
    sheet = workbook["Genome_data"] if "Genome_data" in workbook.sheetnames else workbook.active

    header_row = None
    header_map: Dict[str, int] = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if (cell.value or "").strip().lower() == "biosample_accession":
                header_row = cell.row
                for header_cell in sheet[header_row]:
                    header = header_cell.value
                    if header is None:
                        continue
                    header_str = str(header).strip()
                    if header_str:
                        header_map[header_str] = header_cell.column
                break
        if header_row is not None:
            break

    if header_row is None:
        raise ValueError(f"No header row found in genome template: {template_path}")

    return template_path, sheet, header_row, header_map


def load_sra_instrument_models(workbook: Any) -> set[str]:
    sheet_name = "Library and Platform Terms"
    if sheet_name not in workbook.sheetnames:
        return set()
    sheet = workbook[sheet_name]
    platforms_row = None
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if (cell.value or "").strip() == "Platforms":
                platforms_row = cell.row
                break
        if platforms_row:
            break
    if not platforms_row:
        return set()
    platform_labels = {"ILLUMINA", "PACBIO_SMRT", "OXFORD_NANOPORE"}
    instrument_models: set[str] = set()
    for row in sheet.iter_rows(min_row=platforms_row + 1, max_row=sheet.max_row):
        for col_idx in (3, 7, 10):
            value = sheet.cell(row=row[0].row, column=col_idx).value
            if not value:
                continue
            value_str = str(value).strip()
            if not value_str or value_str in platform_labels:
                continue
            instrument_models.add(value_str)
    return instrument_models


def resolve_edr_path(link: str, edr_root: str) -> Optional[Path]:
    normalized_link = normalize_edr_link(link)
    if not normalized_link or not normalized_link.startswith(EDR_URL_PREFIX):
        return None
    rel_path = normalized_link.replace(EDR_URL_PREFIX, "").lstrip("/")
    edr_root_path = Path(edr_root).resolve()
    return edr_root_path / rel_path


def link_fastq_files(
    genome_data: List[Dict[str, Any]],
    output_dir: str,
    edr_root: str,
    target_subdir: str = "reads_to_upload",
    debug: bool = False,
) -> None:
    target_dir = Path(output_dir) / target_subdir
    target_dir.mkdir(parents=True, exist_ok=True)
    seen_targets: set[str] = set()
    link_count = 0
    for genome in genome_data:
        reads_list = genome.get("reads", []) or []
        for reads in reads_list:
            link = reads.get("link")
            if not link or ".fastq" not in link and ".fq" not in link:
                continue
            edr_path = resolve_edr_path(link, edr_root)
            if not edr_path:
                log_debug(f"Skipping non-EDR link: {link}", enabled=debug)
                continue
            filename = Path(link).name
            target = target_dir / filename
            if target.name in seen_targets:
                continue
            seen_targets.add(target.name)
            if target.exists() or target.is_symlink():
                continue
            try:
                os.symlink(edr_path, target)
                link_count += 1
            except FileExistsError:
                continue
    log_info(f"Symlinked {link_count} file(s) to {target_dir}")


def build_contig_link(genome: Dict[str, Any]) -> str:
    genome_link = genome.get("genome_link") or ""
    if not genome_link:
        return ""
    strain_name = (genome.get("strain") or {}).get("strain_name") or ""
    if strain_name:
        return f"{genome_link.rstrip('/')}/{strain_name}_contigs.fasta"
    return f"{genome_link.rstrip('/')}/contigs.fasta"


def select_contig_link_and_path(
    genome: Dict[str, Any],
    edr_root: str,
    debug: bool = False,
) -> Tuple[str, Optional[Path]]:
    link = build_contig_link(genome)
    if not link:
        return "", None
    edr_path = resolve_edr_path(link, edr_root)
    if not edr_path:
        return link, None
    filename = edr_path.name
    filtered_name = None
    if filename.endswith("_contigs.fasta"):
        filtered_name = filename.replace(
            "_contigs.fasta", "_coverage_filtered_contigs.fasta"
        )
    elif filename == "contigs.fasta":
        filtered_name = "contigs_filtered.fasta"
    if filtered_name:
        filtered_path = edr_path.with_name(filtered_name)
        if filtered_path.exists():
            edr_root_path = Path(edr_root).resolve()
            try:
                rel_path = filtered_path.resolve().relative_to(edr_root_path)
            except ValueError:
                log_debug(
                    f"Filtered contig path outside EDR root: {filtered_path}",
                    enabled=debug,
                )
                return link, edr_path
            filtered_link = f"{EDR_URL_PREFIX.rstrip('/')}/{rel_path.as_posix()}"
            return filtered_link, filtered_path
    return link, edr_path


def infer_assembler_version_from_edr_logs(
    contig_path: Optional[Path],
    assembler_name: Optional[str],
    debug: bool = False,
) -> str:
    if not contig_path or not assembler_name:
        return ""
    if not contig_path.exists():
        return ""
    assembler = assembler_name.strip().lower()
    if assembler not in {"flye", "unicycler"}:
        return ""

    assembly_dir = contig_path.parent
    if not assembly_dir.exists():
        return ""

    log_filename = "flye.log" if assembler == "flye" else "unicycler.log"
    candidate_files: List[Path] = [assembly_dir / log_filename]

    # Expected BERDL layout: logs can also be under results_* folders
    # such as results_flye, results_unicycler, results_circlator_flye.
    for results_dir in sorted(assembly_dir.glob("results_*")):
        if not results_dir.is_dir():
            continue
        candidate_files.append(results_dir / log_filename)
        candidate_files.extend(results_dir.rglob(log_filename))

    candidate_files = sorted(
        {path for path in candidate_files if path.exists()},
        key=lambda path: str(path),
    )
    if not candidate_files:
        log_debug(
            "No assembler logs found for version inference: "
            f"assembler={assembler_name!r} expected={log_filename!r} dir={str(assembly_dir)!r}",
            enabled=debug,
        )
        return ""
    patterns = {
        "flye": [
            re.compile(
                r"\bflye\b[^\n\r]{0,80}?v?(\d+(?:\.\d+){1,3}(?:-[A-Za-z0-9._-]+)?)",
                re.IGNORECASE,
            ),
        ],
        "unicycler": [
            re.compile(
                r"\bunicycler\b[^\n\r]{0,80}?v?(\d+(?:\.\d+){1,3}(?:-[A-Za-z0-9._-]+)?)",
                re.IGNORECASE,
            ),
        ],
    }

    for log_path in candidate_files:
        try:
            text = log_path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        if len(text) > 2_000_000:
            text = text[:2_000_000]
        for pattern in patterns.get(assembler, []):
            match = pattern.search(text)
            if match:
                version = (match.group(1) or "").strip()
                if version:
                    log_debug(
                        "Resolved assembler version from log: "
                        f"assembler={assembler_name!r} version={version!r} file={str(log_path)!r}",
                        enabled=debug,
                    )
                    return version
    return ""


def link_contig_files(
    genome_data: List[Dict[str, Any]],
    output_dir: str,
    edr_root: str,
    target_subdir: str = "contigs_to_upload",
    debug: bool = False,
) -> None:
    target_dir = Path(output_dir) / target_subdir
    target_dir.mkdir(parents=True, exist_ok=True)
    seen_targets: set[str] = set()
    link_count = 0
    for genome in genome_data:
        link, edr_path = select_contig_link_and_path(
            genome, edr_root=edr_root, debug=debug
        )
        if not link or not link.endswith(".fasta"):
            continue
        if not edr_path:
            log_debug(f"Skipping non-EDR link: {link}", enabled=debug)
            continue
        filename = Path(link).name
        target = target_dir / filename
        if target.name in seen_targets:
            continue
        seen_targets.add(target.name)
        if target.exists() or target.is_symlink():
            continue
        try:
            os.symlink(edr_path, target)
            link_count += 1
        except FileExistsError:
            continue
    log_info(f"Symlinked {link_count} file(s) to {target_dir}")


def read_total_bases_from_reads_count(path: Path) -> Optional[int]:
    try:
        with open(path, "r", encoding="utf-8") as handle:
            for line in handle:
                parts = line.rstrip("\n").split("\t", 1)
                if len(parts) == 2 and parts[0] == "total_bases":
                    return int(parts[1])
    except (OSError, ValueError):
        return None
    return None


def reads_count_path_from_reads_file(path: Path) -> Path:
    name = path.name
    for suffix in (".fastq.gz", ".fq.gz", ".fastq", ".fq"):
        if name.endswith(suffix):
            return path.with_name(name[: -len(suffix)] + "_reads_count.txt")
    return path.with_name(name + "_reads_count.txt")


def count_fasta_bases(path: Path) -> Optional[Tuple[int, int, int]]:
    total_chars = 0
    total_lines = 0
    try:
        with open(path, "r", encoding="utf-8") as handle:
            for line in handle:
                if line.startswith(">"):
                    continue
                total_chars += len(line)
                total_lines += 1
    except OSError:
        return None
    total_bases = total_chars - total_lines
    return total_bases, total_chars, total_lines


def compute_coverage_from_files(
    genome: Dict[str, Any],
    edr_root: str,
    debug: bool = False,
) -> Optional[float]:
    contig_link, contig_path = select_contig_link_and_path(
        genome, edr_root=edr_root, debug=debug
    )
    if not contig_path or not contig_path.exists():
        log_debug(
            f"Coverage fallback: contig path missing for {contig_link}",
            enabled=debug,
        )
        return None
    contig_counts = count_fasta_bases(contig_path)
    if not contig_counts:
        log_debug(
            f"Coverage fallback: unable to read contigs at {contig_path}",
            enabled=debug,
        )
        return None
    contig_bases, contig_chars, contig_lines = contig_counts
    if contig_bases <= 0:
        log_debug(
            f"Coverage fallback: contigs bases <= 0 for {contig_path}",
            enabled=debug,
        )
        return None
    reads_total_bases = 0
    reads_files: List[str] = []
    for reads in genome.get("reads", []) or []:
        link = reads.get("link") or ""
        if not link or ".fastq" not in link and ".fq" not in link:
            continue
        edr_path = resolve_edr_path(link, edr_root)
        if not edr_path:
            continue
        count_path = reads_count_path_from_reads_file(edr_path)
        total_bases = read_total_bases_from_reads_count(count_path)
        if total_bases is None:
            continue
        reads_total_bases += total_bases
        reads_files.append(str(count_path))
    if reads_total_bases <= 0:
        log_debug(
            "Coverage fallback: no reads_count files found for coverage estimate.",
            enabled=debug,
        )
        return None
    coverage = reads_total_bases / contig_bases
    log_debug(
        "Coverage fallback: "
        f"reads_total_bases={reads_total_bases} "
        f"contig_bases={contig_bases} "
        f"(chars={contig_chars} lines={contig_lines}) "
        f"coverage={coverage:.4f} "
        f"reads_count_files={reads_files} "
        f"contigs_file={contig_path}",
        enabled=debug,
    )
    return coverage


def normalize_instrument_model(
    instrument_model: Optional[str],
    platform: Optional[str],
    allowed_instruments: set[str],
) -> str:
    if not instrument_model:
        return ""
    if not allowed_instruments:
        return instrument_model
    if instrument_model in allowed_instruments:
        return instrument_model

    lowered = instrument_model.strip().lower()
    allowed_lower = {inst.lower(): inst for inst in allowed_instruments}
    if lowered in allowed_lower:
        return allowed_lower[lowered]

    def strip_vendor(value: str) -> str:
        value = value.strip()
        for prefix in ("illumina ", "pacbio ", "oxford nanopore "):
            if value.lower().startswith(prefix):
                return value[len(prefix) :]
        return value

    stripped = strip_vendor(instrument_model).lower()
    for inst in allowed_instruments:
        if strip_vendor(inst).lower() == stripped:
            return inst

    if platform == "ILLUMINA" and "novaseq" in lowered:
        for candidate in ("Illumina NovaSeq 6000", "NovaSeq 6000"):
            if candidate in allowed_instruments:
                return candidate

    if platform == "ILLUMINA" and not lowered.startswith("illumina "):
        candidate = f"Illumina {instrument_model}"
        if candidate in allowed_instruments:
            return candidate
    if platform == "PACBIO_SMRT" and not lowered.startswith("pacbio "):
        candidate = f"PacBio {instrument_model}"
        if candidate in allowed_instruments:
            return candidate
    if platform == "OXFORD_NANOPORE" and not lowered.startswith("oxford nanopore "):
        candidate = f"Oxford Nanopore {instrument_model}"
        if candidate in allowed_instruments:
            return candidate

    return ""


def get_gtdb_genus_for_strain(
    headers: Dict[str, str],
    strain_name: Optional[str],
    column_cache: Dict[str, List[str]],
) -> Optional[str]:
    if not strain_name:
        return None
    table = "ddt_brick0000522"
    columns = get_table_columns(headers, table, column_cache)
    desired = [
        col
        for col in [
            "sdt_strain_name",
            "taxonomic_level_sys_oterm_name",
            "sdt_taxon_name",
        ]
        if col in columns
    ]
    if "sdt_strain_name" not in desired or "sdt_taxon_name" not in desired:
        return None
    try:
        rows = select_all_rows(
            headers,
            table,
            columns=desired,
            filters=[{"column": "sdt_strain_name", "operator": "=", "value": strain_name}],
        )
    except requests.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 504:
            log_info(f"GTDB genus query timed out for {strain_name}.")
            log_info(
                "Query payload: "
                + str(
                    {
                        "database": "enigma_coral",
                        "table": table,
                        "columns": [{"column": col} for col in desired],
                        "filters": [
                            {
                                "column": "sdt_strain_name",
                                "operator": "=",
                                "value": strain_name,
                            }
                        ],
                        "limit": 1000,
                        "offset": 0,
                    }
                )
            )
            raise
        raise
    genus = None
    family = None
    for row in rows:
        level = (row.get("taxonomic_level_sys_oterm_name") or "").lower()
        if level == "genus":
            genus = row.get("sdt_taxon_name")
        elif level == "family":
            family = row.get("sdt_taxon_name")
    selected = genus or family
    if not selected:
        return None
    if "_" in selected:
        selected = selected.split("_", 1)[0]
    return selected


def build_organism_name(genus: Optional[str], strain_name: Optional[str]) -> Optional[str]:
    if genus and strain_name:
        return f"{genus} sp. {strain_name}"
    if strain_name:
        return strain_name
    return None


def format_lat_lon(location: Dict[str, Any]) -> Optional[str]:
    lat = location.get("latitude")
    lon = location.get("longitude")
    if lat is None or lon is None:
        return None
    lat_dir = "N" if float(lat) >= 0 else "S"
    lon_dir = "E" if float(lon) >= 0 else "W"
    return f"{abs(float(lat))} {lat_dir} {abs(float(lon))} {lon_dir}"


def resolve_collected_by(
    start_token: str, out_lookup: Dict[str, List[Dict[str, Any]]]
) -> Optional[str]:
    visited: set[str] = set()

    def walk(obj_token: str) -> Optional[str]:
        if obj_token in visited:
            return None
        visited.add(obj_token)
        for proc in out_lookup.get(obj_token, []):
            process_name = (proc.get("process_term_name") or "").lower()
            person = (proc.get("person_term_name") or "").strip()
            if "sampling" in process_name and person:
                if "hazen lab" in person.lower():
                    return "Terry Hazen Lab, University of Tennessee, Knoxville, TN, USA"
                return person
            for inp in proc.get("input_objs", []):
                found = walk(inp)
                if found:
                    return found
        return None

    return walk(start_token)


def find_assembly_processes(
    genome_token: str, out_lookup: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    assembly_processes: List[Dict[str, Any]] = []
    visited_procs: set[str] = set()
    visited_objs: set[str] = set()

    def walk(obj: str) -> None:
        if obj in visited_objs:
            return
        visited_objs.add(obj)
        table_name, _ = parse_token(obj)
        for proc in out_lookup.get(obj, []):
            proc_id = proc.get("id")
            if proc_id in visited_procs:
                continue
            visited_procs.add(proc_id)
            process_name = (proc.get("process_term_name") or "").lower()
            if "assembly" in process_name or table_name == "sdt_assembly":
                assembly_processes.append(proc)
            for inp in proc.get("input_objs", []):
                walk(inp)

    walk(genome_token)
    return assembly_processes


def collect_protocols_from_processes(processes: Iterable[Dict[str, Any]]) -> List[str]:
    protocols: List[str] = []
    for proc in processes:
        protocols.extend(normalize_protocol_names(proc.get("protocol")))
    return sorted(set(protocols))


def infer_platform_from_protocols(
    protocol_names: Sequence[str],
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
) -> Tuple[Optional[str], Optional[str]]:
    platform = None
    instrument_model = None
    for name in protocol_names:
        info = get_protocol_info(headers, name, column_cache, protocol_cache)
        seq_tech = (info.get("sequencing_technology") or "").lower()
        machine_type = info.get("machine_type") or ""
        if "illumina" in seq_tech:
            platform = "ILLUMINA"
            instrument_model = machine_type or "Illumina NovaSeq 6000"
            break
        if "nanopore" in seq_tech or "ont" in seq_tech:
            platform = "OXFORD_NANOPORE"
            instrument_model = machine_type or "PromethION"
            break
        if "pacbio" in seq_tech:
            platform = "PACBIO_SMRT"
            instrument_model = machine_type or "PacBio Sequel"
            break
    return platform, instrument_model


def infer_instrument_model_overrides(
    protocol_names: Sequence[str],
    platform: Optional[str],
    sequencing_technology: Optional[str],
    reads_date: Optional[str],
) -> Optional[str]:
    protocol_text = " ".join(protocol_names).lower()
    platform_upper = (platform or "").upper()
    seq_tech_lower = (sequencing_technology or "").lower()

    is_illumina = platform_upper == "ILLUMINA" or "illumina" in seq_tech_lower
    is_ont = platform_upper == "OXFORD_NANOPORE" or "nanopore" in seq_tech_lower or "ont" in seq_tech_lower
    is_pacbio = platform_upper == "PACBIO_SMRT" or "pacbio" in seq_tech_lower

    if "plasmidsaurus" in protocol_text:
        if is_illumina:
            return "NextSeq 2000"
        if is_ont:
            return "PromethION"

    if is_pacbio:
        year = None
        if reads_date:
            match = re.match(r"^(\d{4})-\d{2}-\d{2}$", reads_date)
            if match:
                year = int(match.group(1))
        if year is not None:
            if year >= 2016:
                return "Sequel"
            return "RS II"

    return None


def infer_sra_design_description(
    platform: Optional[str],
    instrument_model: Optional[str],
    protocol_names: Sequence[str],
) -> str:
    platform_upper = (platform or "").upper()
    model_lower = str(instrument_model or "").lower()
    protocol_text = " ".join(str(name or "") for name in protocol_names).lower()

    if platform_upper == "OXFORD_NANOPORE":
        if "plasmidsaurus" in protocol_text:
            return SRA_ONT_PLASMIDSAURUS_DESIGN_DESCRIPTION
        if "lui" in protocol_text or "arkin" in protocol_text:
            return SRA_ONT_LAUREN_DESIGN_DESCRIPTION

    if platform_upper == "ILLUMINA" and any(
        key in model_lower
        for key in (
            "hiseq 4000",
            "novaseq 6000",
            "novaseq x plus",
            "nextseq 2000",
            "nextseq 500",
        )
    ):
        return SRA_ILLUMINA_DESIGN_DESCRIPTION

    if platform_upper == "PACBIO_SMRT" and "sequel" in model_lower:
        return SRA_PACBIO_SEQUEL_DESIGN_DESCRIPTION

    return SRA_DEFAULT_DESIGN_DESCRIPTION


def infer_assembly_method_from_protocols(
    protocol_names: Sequence[str],
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
) -> Optional[str]:
    for name in protocol_names:
        info = get_protocol_info(headers, name, column_cache, protocol_cache)
        desc = (info.get("protocol_description") or "").lower()
        proto_name = (info.get("protocol_name") or name).lower()
        text = f"{proto_name} {desc}"
        # Ambiguous protocol text mentioning multiple assemblers should not
        # be force-resolved to whichever check runs first.
        assembler_hits = 0
        if "metaspades" in text:
            assembler_hits += 1
        if "spades" in text and "metaspades" not in text:
            assembler_hits += 1
        if "flye" in text:
            assembler_hits += 1
        if "canu" in text:
            assembler_hits += 1
        if "unicycler" in text:
            assembler_hits += 1
        if assembler_hits > 1:
            return None
        if "spades" in text:
            version_match = re.search(r"spades\s+v?(\d+\.\d+\.\d+)", text)
            return f"SPAdes {version_match.group(1)}" if version_match else "SPAdes"
        if "flye" in text:
            # Prefer explicit Flye tool version (often in parentheses) over wrapper versions such as kb_flye.
            version_match = re.search(
                r"\(\s*flye\s+v?(\d+(?:\.\d+){1,3})\s*\)", text, re.IGNORECASE
            )
            if not version_match:
                version_match = re.search(
                    r"\bflye\b\s+v?(\d+(?:\.\d+){1,3})", text, re.IGNORECASE
                )
            return f"Flye {version_match.group(1)}" if version_match else "Flye"
        if "canu" in text:
            version_match = re.search(r"canu\s+v?(\d+\.\d+)", text)
            return f"CANU {version_match.group(1)}" if version_match else "CANU"
        if "metaspades" in text:
            return "MetaSPAdes"
        if "unicycler" in text:
            return "Unicycler"
    return None


def split_assembly_method(method: Optional[str], debug: bool = False) -> Tuple[str, str]:
    if not method:
        return "", ""
    text = method.strip()
    match = re.search(r"(?:v\.?\s*)?(\d+(?:\.\d+){1,3})", text, re.IGNORECASE)
    if not match:
        if debug:
            log_debug(
                f"Assembly method parse: raw={method!r} -> method={text!r} version=''",
                enabled=debug,
            )
        return text, ""
    version = match.group(1)
    name = (text[: match.start()] + text[match.end() :]).strip(" -_()")
    if not name:
        name = text.strip()
    if debug:
        log_debug(
            f"Assembly method parse: raw={method!r} -> method={name!r} version={version!r}",
            enabled=debug,
        )
    return name, version


def normalize_assembler_version(version: Optional[str]) -> str:
    if version is None:
        return ""
    text = str(version).strip()
    if not text:
        return ""
    if text.lower() == "unknown":
        return "unknown"
    return re.sub(r"^[vV]\s*", "", text)


def choose_assembly_date(processes: Iterable[Dict[str, Any]]) -> str:
    dates = [proc.get("date_end") for proc in processes if proc.get("date_end")]
    if not dates:
        return ""
    return sorted(dates)[-1]


def normalize_strain_name(genome_name: str, strain: Optional[Dict[str, Any]]) -> str:
    if strain and strain.get("strain_name"):
        return str(strain.get("strain_name"))
    if genome_name:
        return genome_name.split(".", 1)[0]
    return ""


def fetch_read_coverage_map(
    headers: Dict[str, str],
    strain_names: Sequence[str],
    column_cache: Dict[str, List[str]],
) -> Dict[str, float]:
    if not strain_names:
        return {}
    columns = get_table_columns(headers, READ_COVERAGE_TABLE, column_cache)
    if "sdt_strain_name" not in columns or READ_COVERAGE_COLUMN not in columns:
        return {}
    rows = select_all_rows(
        headers,
        READ_COVERAGE_TABLE,
        columns=["sdt_strain_name", READ_COVERAGE_COLUMN],
        filters=None,
        limit=1000,
    )
    coverage_values: Dict[str, List[float]] = defaultdict(list)
    for row in rows:
        strain = row.get("sdt_strain_name")
        value = row.get(READ_COVERAGE_COLUMN)
        if strain is None or value is None:
            continue
        if strain_names and str(strain) not in strain_names:
            continue
        try:
            coverage_values[str(strain)].append(float(value))
        except (TypeError, ValueError):
            continue
    coverage_map: Dict[str, float] = {}
    for strain, values in coverage_values.items():
        if values:
            coverage_map[strain] = sum(values) / len(values)
    return coverage_map


def _parse_accession_key(accession: str) -> Tuple[int, int, int]:
    match = re.match(r"^(GC[AF])_(\d+)(?:\.(\d+))?$", accession.strip(), re.IGNORECASE)
    if not match:
        return (0, -1, -1)
    prefix = match.group(1).upper()
    number = int(match.group(2))
    version = int(match.group(3) or 0)
    prefix_rank = 2 if prefix == "GCF" else 1
    return (prefix_rank, number, version)


def _choose_refseq_accession(values: Sequence[str]) -> str:
    candidates = [
        value.strip()
        for value in values
        if isinstance(value, str)
        and re.match(r"^GC[AF]_\d+(?:\.\d+)?$", value.strip(), flags=re.IGNORECASE)
    ]
    if not candidates:
        return ""
    unique_candidates = sorted(set(candidates), key=_parse_accession_key, reverse=True)
    return unique_candidates[0]


def fetch_isolate_refseq_map(
    headers: Dict[str, str],
    strain_names: Sequence[str],
    discovered_tables: Sequence[str],
    column_cache: Dict[str, List[str]],
) -> Dict[str, str]:
    unique_strains = sorted({name for name in strain_names if name})
    if not unique_strains:
        return {}

    table_name = GENBANK_LINK_TABLE
    required_columns = {
        "sdt_strain_name",
        "link_sequence_type_genome_sequence_database_genbank",
    }
    discovered_normalized = {str(name).strip().lower() for name in discovered_tables}
    if table_name.lower() not in discovered_normalized:
        log_info(
            f"Current GenBank link table {table_name} not listed by table discovery; "
            "attempting direct schema lookup."
        )
    try:
        columns_in_table = set(get_table_columns(headers, table_name, column_cache))
    except requests.HTTPError:
        log_info(
            f"Current GenBank link table {table_name} unavailable via schema lookup; "
            "leaving update_for blank."
        )
        return {}
    if not required_columns.issubset(columns_in_table):
        log_info(
            f"Current GenBank link table {table_name} missing required columns; leaving update_for blank."
        )
        return {}
    log_info(f"Using {table_name} for isolate RefSeq lookup")
    refseq_values: Dict[str, List[str]] = defaultdict(list)
    chunk_size = 200
    columns = [
        "sdt_strain_name",
        "link_sequence_type_genome_sequence_database_genbank",
    ]
    for idx in range(0, len(unique_strains), chunk_size):
        chunk = unique_strains[idx : idx + chunk_size]
        try:
            rows = select_all_rows(
                headers,
                table_name,
                columns=columns,
                filters=[{"column": "sdt_strain_name", "operator": "IN", "value": chunk}],
                limit=1000,
            )
        except requests.HTTPError as exc:
            if exc.response is None or exc.response.status_code != 400:
                raise
            # Some deployments reject IN filters for this table; fall back to
            # one equality query per strain.
            rows = []
            for strain_name in chunk:
                rows.extend(
                    select_all_rows(
                        headers,
                        table_name,
                        columns=columns,
                        filters=[
                            {
                                "column": "sdt_strain_name",
                                "operator": "=",
                                "value": strain_name,
                            }
                        ],
                        limit=1000,
                    )
                )
        for row in rows:
            strain_name = row.get("sdt_strain_name")
            accession = row.get("link_sequence_type_genome_sequence_database_genbank")
            if not strain_name or not accession:
                continue
            refseq_values[str(strain_name)].append(str(accession))

    refseq_map: Dict[str, str] = {}
    for strain_name, values in refseq_values.items():
        chosen = _choose_refseq_accession(values)
        if chosen:
            refseq_map[strain_name] = chosen
    return refseq_map


def infer_sequencing_techs(
    protocol_names: Sequence[str],
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
) -> List[str]:
    techs: set[str] = set()
    for name in protocol_names:
        info = get_protocol_info(headers, name, column_cache, protocol_cache)
        tech = info.get("sequencing_technology")
        if tech:
            techs.add(tech)
    return sorted(techs)


def is_long_read_assembly_from_protocols(
    protocol_names: Sequence[str],
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
) -> bool:
    for tech in infer_sequencing_techs(protocol_names, headers, column_cache, protocol_cache):
        if is_long_read_tech(None, tech):
            return True
    return False


def is_long_read_submission_context(
    protocol_names: Sequence[str],
    reads_list: Sequence[Dict[str, Any]],
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
) -> bool:
    protocol_techs = infer_sequencing_techs(
        protocol_names, headers, column_cache, protocol_cache
    )
    if protocol_techs:
        return any(is_long_read_tech(None, tech) for tech in protocol_techs)
    for reads in reads_list or []:
        seq_tech = (
            reads.get("source_sequencing_technology")
            or reads.get("sequencing_technology")
            or ""
        )
        platform = ""
        if "pacbio" in str(seq_tech).lower():
            platform = "PACBIO_SMRT"
        elif "nanopore" in str(seq_tech).lower() or "ont" in str(seq_tech).lower():
            platform = "OXFORD_NANOPORE"
        if is_long_read_tech(platform, str(seq_tech)):
            return True
    return False


def is_hybrid_reads_context(reads_list: Sequence[Dict[str, Any]]) -> bool:
    has_long = False
    has_illumina = False
    for reads in reads_list or []:
        seq_tech = (
            reads.get("source_sequencing_technology")
            or reads.get("sequencing_technology")
            or ""
        )
        seq_lower = str(seq_tech).lower()
        if "illumina" in seq_lower:
            has_illumina = True
        if is_long_read_tech(None, seq_tech):
            has_long = True
    return has_long and has_illumina


def infer_plasmidsaurus_protocol_year(protocol_names: Sequence[str]) -> Optional[str]:
    for name in protocol_names:
        text = str(name or "").lower()
        if "plasmidsaurus" not in text:
            continue
        match = re.search(r"(20\d{2})", text)
        if match:
            return match.group(1)
    return None


def is_plasmidsaurus_protocol(protocol_names: Sequence[str]) -> bool:
    return infer_plasmidsaurus_protocol_year(protocol_names) is not None


def has_long_reads_for_year(reads_list: Sequence[Dict[str, Any]], year: str) -> bool:
    year_prefix = f"{year}-"
    for reads in reads_list or []:
        seq_tech = (
            reads.get("source_sequencing_technology")
            or reads.get("sequencing_technology")
            or ""
        )
        if not is_long_read_tech(None, seq_tech):
            continue
        link_name = Path(str(reads.get("link") or "")).name
        reads_name = str(reads.get("reads_name") or "")
        reads_date = extract_date_from_filenames(link_name, reads_name)
        if reads_date and reads_date.startswith(year_prefix):
            return True
    return False


def infer_isolation_source(sample_name: str, sample: Dict[str, Any]) -> Optional[str]:
    material = sample.get("material_name") or ""
    description = sample.get("description") or ""
    text = " ".join(part for part in [material, description, sample_name] if part).lower()
    if (
        "groundwater" in text
        or re.search(r"\bground\s*water\b", text)
        or "subsurface water" in text
        or "aquifer" in text
    ):
        return f"Groundwater sample {sample_name} from ORR".strip()
    if (
        re.search(r"\bsoil\b", text)
        or "topsoil" in text
        or "subsoil" in text
        or "rhizosphere" in text
        or "sediment" in text
    ):
        if "sediment" in text:
            return f"Sediment sample {sample_name} from ORR".strip()
        return f"Soil sample {sample_name} from ORR".strip()
    return None


def format_depth_meters(depth_value: Any) -> str:
    if depth_value in (None, ""):
        return ""
    try:
        depth_float = float(depth_value)
    except (TypeError, ValueError):
        depth_str = str(depth_value).strip()
        if not depth_str:
            return ""
        return f"{depth_str} m"
    if depth_float.is_integer():
        depth_str = str(int(depth_float))
    else:
        depth_str = str(depth_float)
    return f"{depth_str} m"


def build_biosample_name(sample_name: str, strain_name: str) -> str:
    formatted = f"environmental sample {sample_name} isolate {strain_name}".strip()
    return formatted


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_date_for_compare(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    match = re.match(r"^(\d{4}-\d{2}-\d{2})", text)
    if match:
        return match.group(1)
    return text


def _to_float(value: Any) -> Optional[float]:
    text = _normalize_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _values_match(
    berdl_value: Any,
    metadata_value: Any,
    *,
    numeric_tolerance: Optional[float] = None,
    normalize_date: bool = False,
) -> bool:
    if normalize_date:
        left = _normalize_date_for_compare(berdl_value)
        right = _normalize_date_for_compare(metadata_value)
        return bool(left and right and left == right)

    if numeric_tolerance is not None:
        left_num = _to_float(berdl_value)
        right_num = _to_float(metadata_value)
        if left_num is not None and right_num is not None:
            return abs(left_num - right_num) <= numeric_tolerance

    left_text = _normalize_text(berdl_value)
    right_text = _normalize_text(metadata_value)
    return bool(left_text and right_text and left_text.casefold() == right_text.casefold())


def should_promote_comment_to_description(comment: str) -> bool:
    text = _normalize_text(comment).lower()
    if not text:
        return False
    keywords = [
        "capillary fringe",
        "vadose",
        "variably saturated",
        "saturated",
        "unsaturated",
        "clay",
        "sand",
        "silt",
        "gravel",
        "brown",
        "red",
        "gray",
        "grey",
        "black",
        "orange",
        "yellow",
        "green",
        "blue",
        "wet",
        "dry",
        "water",
        "groundwater",
    ]
    return any(keyword in text for keyword in keywords)


def format_biosample_description_from_comment(comment: str) -> str:
    text = _normalize_text(comment)
    if not text:
        return ""
    lowered = text.lower()
    blocked_prefixes = ("ground water", "groundwater", "3l", "3 l", "initial", "unfiltered")
    if lowered.startswith(blocked_prefixes):
        return ""
    if not should_promote_comment_to_description(text):
        return ""
    return f"soil sample: {text}"


def load_sample_metadata(
    metadata_path: Optional[str], debug: bool = False
) -> Dict[str, Dict[str, str]]:
    path_candidates: List[Path] = []
    if metadata_path:
        path_candidates.append(Path(metadata_path))
    else:
        path_candidates.extend(SAMPLE_METADATA_CANDIDATES)
    metadata_file = next((path for path in path_candidates if path.exists()), None)
    if not metadata_file:
        log_info("No sample_metadata.tsv found; skipping metadata enrichment.")
        return {}

    log_info(f"Loading sample metadata from {metadata_file}")
    rows_by_sample_id: Dict[str, Dict[str, str]] = {}
    with open(metadata_file, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            if not row:
                continue
            sample_id = _normalize_text(row.get("sample ID"))
            if not sample_id:
                continue
            rows_by_sample_id[sample_id] = {k: _normalize_text(v) for k, v in row.items() if k}
    log_info(f"Loaded {len(rows_by_sample_id)} sample metadata row(s)")
    if debug and rows_by_sample_id:
        sample_keys = list(rows_by_sample_id.keys())[:5]
        log_debug(f"sample_metadata.tsv sample IDs (first 5): {sample_keys}", enabled=True)
    return rows_by_sample_id


def validate_sample_metadata_match(
    genome_name: str,
    sample: Dict[str, Any],
    location: Dict[str, Any],
    metadata: Dict[str, str],
) -> List[str]:
    mismatches: List[str] = []
    checks = [
        (
            "sampling_date",
            sample.get("date"),
            metadata.get("sampling date"),
            {"normalize_date": True},
        ),
        (
            "location/well ID",
            sample.get("location_name"),
            metadata.get("location/well ID"),
            {},
        ),
        (
            "latitude",
            location.get("latitude"),
            metadata.get("latitude"),
            {"numeric_tolerance": 1e-6},
        ),
        (
            "longitude",
            location.get("longitude"),
            metadata.get("longitude"),
            {"numeric_tolerance": 1e-6},
        ),
        (
            "material",
            sample.get("material_name"),
            metadata.get("material"),
            {},
        ),
        (
            "sampling depth",
            sample.get("depth_meter"),
            metadata.get("mean sampling depth (m)"),
            {"numeric_tolerance": 1e-6},
        ),
    ]
    for field_name, berdl_value, metadata_value, options in checks:
        metadata_text = _normalize_text(metadata_value)
        berdl_text = _normalize_text(berdl_value)
        if not metadata_text:
            continue
        if not berdl_text:
            mismatches.append(
                f"{genome_name}: sample metadata has {field_name}={metadata_value!r} "
                "but BERDL value is missing"
            )
            continue
        if not _values_match(berdl_value, metadata_value, **options):
            mismatches.append(
                f"{genome_name}: sample metadata mismatch for {field_name}: "
                f"BERDL={berdl_value!r} sample_metadata.tsv={metadata_value!r}"
            )
    return mismatches


def summarize_sample_metadata_match(
    sample: Dict[str, Any], location: Dict[str, Any], metadata: Dict[str, str]
) -> Dict[str, Any]:
    checks = [
        (
            "sampling_date",
            sample.get("date"),
            metadata.get("sampling date"),
            {"normalize_date": True},
        ),
        (
            "location/well ID",
            sample.get("location_name"),
            metadata.get("location/well ID"),
            {},
        ),
        (
            "latitude",
            location.get("latitude"),
            metadata.get("latitude"),
            {"numeric_tolerance": 1e-6},
        ),
        (
            "longitude",
            location.get("longitude"),
            metadata.get("longitude"),
            {"numeric_tolerance": 1e-6},
        ),
        (
            "material",
            sample.get("material_name"),
            metadata.get("material"),
            {},
        ),
        (
            "sampling depth",
            sample.get("depth_meter"),
            metadata.get("mean sampling depth (m)"),
            {"numeric_tolerance": 1e-6},
        ),
    ]

    compared_fields: List[str] = []
    matched_fields: List[str] = []
    mismatched_fields: List[str] = []

    for field_name, berdl_value, metadata_value, options in checks:
        metadata_text = _normalize_text(metadata_value)
        if not metadata_text:
            continue
        compared_fields.append(field_name)
        berdl_text = _normalize_text(berdl_value)
        if berdl_text and _values_match(berdl_value, metadata_value, **options):
            matched_fields.append(field_name)
        else:
            mismatched_fields.append(field_name)

    return {
        "compared_count": len(compared_fields),
        "matched_count": len(matched_fields),
        "compared_fields": compared_fields,
        "matched_fields": matched_fields,
        "mismatched_fields": mismatched_fields,
    }


def ensure_header_column(
    sheet: Any, header_row: int, header_map: Dict[str, int], header_name: str
) -> int:
    existing = header_map.get(header_name)
    if existing is not None:
        return existing
    max_used_col = 0
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=header_row, column=col).value not in (None, ""):
            max_used_col = col
    new_col = max_used_col + 1 if max_used_col else sheet.max_column + 1
    sheet.cell(row=header_row, column=new_col, value=header_name)
    header_map[header_name] = new_col
    return new_col


def extract_date_from_filenames(*filenames: Optional[str]) -> Optional[str]:
    for filename in filenames:
        if not filename:
            continue
        match = re.search(r"\d{4}-\d{2}-\d{2}", filename)
        if match:
            return match.group(0)
    return None


def infer_replicate_number_from_filenames(*filenames: Optional[str]) -> Optional[int]:
    pattern = re.compile(
        r"(?:^|[_-])rep(?:licate)?\s*0*(\d+)(?=[_.-]|$)", re.IGNORECASE
    )
    for filename in filenames:
        if not filename:
            continue
        match = pattern.search(str(filename))
        if not match:
            continue
        try:
            value = int(match.group(1))
        except (TypeError, ValueError):
            continue
        if value > 0:
            return value
    return None


def infer_read_tech_label(platform: Optional[str], sequencing_tech: Optional[str]) -> str:
    platform_upper = (platform or "").upper()
    seq_lower = (sequencing_tech or "").lower()
    if "PACBIO" in platform_upper or "pacbio" in seq_lower:
        return "Pacbio"
    if "NANOPORE" in platform_upper or "nanopore" in seq_lower or "ont" in seq_lower:
        return "Nanopore"
    return "Illumina"


def is_long_read_tech(platform: Optional[str], sequencing_tech: Optional[str]) -> bool:
    platform_upper = (platform or "").upper()
    seq_lower = (sequencing_tech or "").lower()
    return bool(
        "PACBIO" in platform_upper
        or "NANOPORE" in platform_upper
        or "pacbio" in seq_lower
        or "nanopore" in seq_lower
        or "ont" in seq_lower
    )


def generate_biosample_table(
    genome_data: List[Dict[str, Any]],
    output_file: str,
    output_dir: str,
    debug: bool = False,
    dry_run: bool = False,
) -> int:
    template_path, sheet, header_row, header_map = load_biosample_template_workbook(
        output_dir
    )

    last_data_row = header_row
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        if any(cell.value not in (None, "") for cell in row):
            last_data_row = row[0].row

    next_row = last_data_row + 1
    start_row = next_row
    custom_headers = ["Moisture (%)", "Conductivity (mS/cm)", "pH"]
    for custom_header in custom_headers:
        ensure_header_column(sheet, header_row, header_map, custom_header)

    for genome in genome_data:
        sample = genome.get("sample", {}) or {}
        location = genome.get("location", {}) or {}
        strain = genome.get("strain", {}) or {}
        sample_metadata = genome.get("sample_metadata", {}) or {}

        sample_name = sample.get("sample_name") or ""
        strain_name = strain.get("strain_name") or sample.get("sample_name")
        formatted_sample_name = build_biosample_name(sample_name, strain_name)
        genus = genome.get("gtdb_genus")
        organism_name = build_organism_name(genus, strain_name)

        isolation_source = infer_isolation_source(sample_name, sample)
        metadata_comment = sample_metadata.get("comments", "")
        description_value = format_biosample_description_from_comment(metadata_comment)
        if debug:
            log_debug(
                "Biosample isolation_source debug: "
                f"sample_name={sample_name!r} "
                f"material_name={(sample.get('material_name') or '')!r} "
                f"description={(sample.get('description') or '')!r} "
                f"resolved={isolation_source!r}",
                enabled=True,
            )

        values = {
            "*sample_name": formatted_sample_name,
            "*organism": organism_name or "",
            "strain": strain_name or "",
            "isolate": "",
            "host": "",
            "isolation_source": isolation_source or "",
            "*collection_date": sample.get("date") or "",
            "*geo_loc_name": "USA: Tennessee, Oak Ridge Reservation (ORR)",
            "*sample_type": "cell culture",
            "collected_by": BIOSAMPLE_COLLECTED_BY_OVERRIDE,
            "depth": format_depth_meters(sample.get("depth_meter")),
            "env_broad_scale": "temperate woodland biome [ENVO:01000221]",
            "lat_lon": format_lat_lon(location) or "",
            "temp": sample_metadata.get("Temperature (Celsius)") or "",
            "description": description_value,
            "Moisture (%)": sample_metadata.get("Moisture (%)") or "",
            "Conductivity (mS/cm)": sample_metadata.get("Conductivity (mS/cm)") or "",
            "pH": sample_metadata.get("pH") or "",
        }

        for header, value in values.items():
            col = header_map.get(header)
            if col is not None:
                sheet.cell(row=next_row, column=col, value=value)
        next_row += 1

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not dry_run:
        sheet.parent.save(output_path)
    return max(0, next_row - start_row)


def generate_sra_table(
    genome_data: List[Dict[str, Any]],
    output_file: str,
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
    debug: bool = False,
    dry_run: bool = False,
) -> int:
    template_path, sheet, header_row, header_map = load_sra_template_workbook(
        Path(output_file).parent
    )
    allowed_instruments = load_sra_instrument_models(sheet.parent)
    if debug:
        log_debug(
            "SRA template headers: "
            f"platform_col={header_map.get('platform')} "
            f"instrument_model_col={header_map.get('instrument_model')} "
            f"allowed_instruments={len(allowed_instruments)}",
            enabled=True,
        )

    last_data_row = header_row
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        if any(cell.value not in (None, "") for cell in row):
            last_data_row = row[0].row

    next_row = last_data_row + 1
    start_row = next_row
    used_library_ids: set[str] = set()
    next_suffix_by_library_stem: Dict[str, int] = {}
    short_read_date_flag: Dict[str, bool] = {}
    short_read_dates: Dict[str, List[Optional[str]]] = {}
    for genome in genome_data:
        strain = genome.get("strain", {}) or {}
        sample = genome.get("sample", {}) or {}
        sample_name = sample.get("sample_name") or ""
        isolate_name = strain.get("strain_name") or sample.get("sample_name") or genome.get(
            "genome_name", ""
        )
        assembly_protocols = collect_protocols_from_processes(
            genome.get("assembly_processes", []) or []
        )
        plasmidsaurus_year = infer_plasmidsaurus_protocol_year(assembly_protocols)
        prefer_plasmidsaurus_year_long_reads = bool(
            plasmidsaurus_year
            and has_long_reads_for_year(genome.get("reads", []) or [], plasmidsaurus_year)
        )
        use_long_read_hardcoded_metadata = is_long_read_submission_context(
            assembly_protocols,
            genome.get("reads", []) or [],
            headers,
            column_cache,
            protocol_cache,
        )
        hardcoded_meta = (
            hardcoded_metadata_for_isolate(isolate_name)
            if use_long_read_hardcoded_metadata
            else {}
        )
        hardcoded_illumina = hardcoded_meta.get("illumina", "")
        hybrid_reads_context = is_hybrid_reads_context(genome.get("reads", []) or [])
        biosample_name = build_biosample_name(sample_name, isolate_name)
        reads_list = genome.get("reads", []) or []

        reads_by_base_name: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for reads in reads_list:
            reads_name = reads.get("reads_name", "") or ""
            base_name = re.sub(
                r"[_-](R[12]|fwd|rev|forward|reverse|paired|unpaired).*$",
                "",
                reads_name,
                flags=re.IGNORECASE,
            )
            reads_by_base_name[base_name].append(reads)

        for base_name, reads_group in reads_by_base_name.items():
            if len(reads_group) >= 2:
                library_layout = "PAIRED"
                forward_reads = None
                reverse_reads = None
                for reads in reads_group:
                    name_lower = (reads.get("reads_name") or "").lower()
                    if any(
                        x in name_lower
                        for x in ["_r1", "_fwd", "_forward", "_1.fastq", "_1.fq"]
                    ):
                        forward_reads = reads
                    elif any(
                        x in name_lower
                        for x in ["_r2", "_rev", "_reverse", "_2.fastq", "_2.fq"]
                    ):
                        reverse_reads = reads
                if not forward_reads:
                    forward_reads = reads_group[0]
                if not reverse_reads and len(reads_group) > 1:
                    reverse_reads = reads_group[1]
                filename = (forward_reads.get("link") or "").split("/")[-1]
                filename2 = (reverse_reads.get("link") or "").split("/")[-1]
                primary_reads = forward_reads
            else:
                primary_reads = reads_group[0] if reads_group else None
                filename = (
                    (primary_reads.get("link") or "").split("/")[-1]
                    if primary_reads
                    else ""
                )
                filename2 = None
                if primary_reads and is_paired_read_type(primary_reads.get("read_type")):
                    library_layout = "PAIRED"
                else:
                    library_layout = "SINGLE"

            if not primary_reads:
                continue

            source_protocols = primary_reads.get("source_protocols", []) or []
            reads_date = extract_date_from_filenames(filename, filename2)
            seq_tech_debug = (
                primary_reads.get("source_sequencing_technology")
                or primary_reads.get("sequencing_technology")
                or ""
            )
            if source_protocols:
                protocol_names = list(source_protocols)
            else:
                protocol_names = []
                protocol_names.extend(normalize_protocol_names(sample.get("protocol")))
                protocol_names.extend(primary_reads.get("protocols", []))
                if not protocol_names:
                    assembly_protocols = collect_protocols_from_processes(
                        genome.get("assembly_processes", []) or []
                    )
                    protocol_names.extend(assembly_protocols)
            platform, instrument_model = infer_platform_from_protocols(
                protocol_names, headers, column_cache, protocol_cache
            )

            if not platform:
                seq_tech = (
                    primary_reads.get("source_sequencing_technology")
                    or primary_reads.get("sequencing_technology")
                    or ""
                ).lower()
                if "illumina" in seq_tech:
                    platform = "ILLUMINA"
                    instrument_model = "unknown"
                elif "nanopore" in seq_tech or "ont" in seq_tech:
                    platform = "OXFORD_NANOPORE"
                    instrument_model = "unknown"
                elif "pacbio" in seq_tech:
                    platform = "PACBIO_SMRT"
                    instrument_model = "unknown"
                else:
                    platform = "ILLUMINA"
                    instrument_model = "unknown"

            seq_tech = (
                primary_reads.get("source_sequencing_technology")
                or primary_reads.get("sequencing_technology")
                or ""
            )
            override_model = infer_instrument_model_overrides(
                protocol_names, platform, seq_tech, reads_date
            )
            if override_model:
                instrument_model = override_model
            enforce_illumina_match(
                normalize_isolate_key(isolate_name),
                platform,
                instrument_model,
                hardcoded_illumina,
            )
            if (
                use_long_read_hardcoded_metadata
                and hybrid_reads_context
                and "illumina" in str(seq_tech).lower()
                and _illumina_family(hardcoded_illumina) == "novaseq"
            ):
                platform = "ILLUMINA"
            hardcoded_model = hardcoded_illumina_model(hardcoded_illumina)
            if platform == "ILLUMINA" and hardcoded_model:
                instrument_model = hardcoded_model

            if debug:
                log_debug(
                    "SRA instrument debug: "
                    f"sample={biosample_name!r} "
                    f"reads={primary_reads.get('reads_name')!r} "
                    f"source_reads={primary_reads.get('source_reads_token')!r} "
                    f"source_protocols={source_protocols!r} "
                    f"protocol_names={protocol_names!r} "
                    f"seq_tech={seq_tech_debug!r} "
                    f"platform={platform!r} "
                    f"instrument_model={instrument_model!r}",
                    enabled=True,
                )

            instrument_model = normalize_instrument_model(
                instrument_model if instrument_model != "unknown" else "",
                platform,
                allowed_instruments,
            )
            if debug and instrument_model == "":
                log_debug(
                    "SRA instrument filtered: "
                    f"instrument_model not in template list",
                    enabled=True,
                )

            seq_tech = (
                primary_reads.get("source_sequencing_technology")
                or primary_reads.get("sequencing_technology")
            )
            tech_label = infer_read_tech_label(platform, seq_tech)
            long_read = is_long_read_tech(platform, seq_tech)
            if (
                plasmidsaurus_year
                and prefer_plasmidsaurus_year_long_reads
                and long_read
                and reads_date
                and not reads_date.startswith(f"{plasmidsaurus_year}-")
            ):
                if debug:
                    log_debug(
                        "Skipping long read due to plasmidsaurus protocol year mismatch: "
                        f"isolate={isolate_name!r} reads={primary_reads.get('reads_name')!r} "
                        f"reads_date={reads_date!r} protocol_year={plasmidsaurus_year!r}",
                        enabled=True,
                    )
                continue

            if not long_read:
                short_read_dates.setdefault(isolate_name, []).append(reads_date)

    for isolate_name, dates in short_read_dates.items():
        if len(dates) > 1 and any(date for date in dates):
            short_read_date_flag[isolate_name] = True

    for genome in genome_data:
        strain = genome.get("strain", {}) or {}
        sample = genome.get("sample", {}) or {}
        sample_name = sample.get("sample_name") or ""
        isolate_name = strain.get("strain_name") or sample.get("sample_name") or genome.get(
            "genome_name", ""
        )
        assembly_protocols = collect_protocols_from_processes(
            genome.get("assembly_processes", []) or []
        )
        plasmidsaurus_year = infer_plasmidsaurus_protocol_year(assembly_protocols)
        prefer_plasmidsaurus_year_long_reads = bool(
            plasmidsaurus_year
            and has_long_reads_for_year(genome.get("reads", []) or [], plasmidsaurus_year)
        )
        use_long_read_hardcoded_metadata = is_long_read_submission_context(
            assembly_protocols,
            genome.get("reads", []) or [],
            headers,
            column_cache,
            protocol_cache,
        )
        hardcoded_meta = (
            hardcoded_metadata_for_isolate(isolate_name)
            if use_long_read_hardcoded_metadata
            else {}
        )
        hardcoded_illumina = hardcoded_meta.get("illumina", "")
        hybrid_reads_context = is_hybrid_reads_context(genome.get("reads", []) or [])
        biosample_name = build_biosample_name(sample_name, isolate_name)
        reads_list = genome.get("reads", []) or []

        reads_by_base_name = defaultdict(list)
        for reads in reads_list:
            reads_name = reads.get("reads_name", "") or ""
            base_name = re.sub(
                r"[_-](R[12]|fwd|rev|forward|reverse|paired|unpaired).*$",
                "",
                reads_name,
                flags=re.IGNORECASE,
            )
            reads_by_base_name[base_name].append(reads)

        for base_name, reads_group in reads_by_base_name.items():
            if len(reads_group) >= 2:
                library_layout = "PAIRED"
                forward_reads = None
                reverse_reads = None
                for reads in reads_group:
                    name_lower = (reads.get("reads_name") or "").lower()
                    if any(
                        x in name_lower
                        for x in ["_r1", "_fwd", "_forward", "_1.fastq", "_1.fq"]
                    ):
                        forward_reads = reads
                    elif any(
                        x in name_lower
                        for x in ["_r2", "_rev", "_reverse", "_2.fastq", "_2.fq"]
                    ):
                        reverse_reads = reads
                if not forward_reads:
                    forward_reads = reads_group[0]
                if not reverse_reads and len(reads_group) > 1:
                    reverse_reads = reads_group[1]
                filename = (forward_reads.get("link") or "").split("/")[-1]
                filename2 = (reverse_reads.get("link") or "").split("/")[-1]
                primary_reads = forward_reads
            else:
                primary_reads = reads_group[0] if reads_group else None
                filename = (
                    (primary_reads.get("link") or "").split("/")[-1]
                    if primary_reads
                    else ""
                )
                filename2 = None
                if primary_reads and is_paired_read_type(primary_reads.get("read_type")):
                    library_layout = "PAIRED"
                else:
                    library_layout = "SINGLE"

            if not primary_reads:
                continue

            source_protocols = primary_reads.get("source_protocols", []) or []
            reads_date = extract_date_from_filenames(filename, filename2)
            seq_tech_debug = (
                primary_reads.get("source_sequencing_technology")
                or primary_reads.get("sequencing_technology")
                or ""
            )
            if source_protocols:
                protocol_names = list(source_protocols)
            else:
                protocol_names = []
                protocol_names.extend(normalize_protocol_names(sample.get("protocol")))
                protocol_names.extend(primary_reads.get("protocols", []))
                if not protocol_names:
                    assembly_protocols = collect_protocols_from_processes(
                        genome.get("assembly_processes", []) or []
                    )
                    protocol_names.extend(assembly_protocols)
            platform, instrument_model = infer_platform_from_protocols(
                protocol_names, headers, column_cache, protocol_cache
            )

            if not platform:
                seq_tech = (
                    primary_reads.get("source_sequencing_technology")
                    or primary_reads.get("sequencing_technology")
                    or ""
                ).lower()
                if "illumina" in seq_tech:
                    platform = "ILLUMINA"
                    instrument_model = "unknown"
                elif "nanopore" in seq_tech or "ont" in seq_tech:
                    platform = "OXFORD_NANOPORE"
                    instrument_model = "unknown"
                elif "pacbio" in seq_tech:
                    platform = "PACBIO_SMRT"
                    instrument_model = "unknown"
                else:
                    platform = "ILLUMINA"
                    instrument_model = "unknown"

            seq_tech = (
                primary_reads.get("source_sequencing_technology")
                or primary_reads.get("sequencing_technology")
                or ""
            )
            override_model = infer_instrument_model_overrides(
                protocol_names, platform, seq_tech, reads_date
            )
            if override_model:
                instrument_model = override_model
            enforce_illumina_match(
                normalize_isolate_key(isolate_name),
                platform,
                instrument_model,
                hardcoded_illumina,
            )
            if (
                use_long_read_hardcoded_metadata
                and hybrid_reads_context
                and "illumina" in str(seq_tech).lower()
                and _illumina_family(hardcoded_illumina) == "novaseq"
            ):
                platform = "ILLUMINA"
            hardcoded_model = hardcoded_illumina_model(hardcoded_illumina)
            if platform == "ILLUMINA" and hardcoded_model:
                instrument_model = hardcoded_model

            if debug:
                log_debug(
                    "SRA instrument debug: "
                    f"sample={biosample_name!r} "
                    f"reads={primary_reads.get('reads_name')!r} "
                    f"source_reads={primary_reads.get('source_reads_token')!r} "
                    f"source_protocols={source_protocols!r} "
                    f"protocol_names={protocol_names!r} "
                    f"seq_tech={seq_tech_debug!r} "
                    f"platform={platform!r} "
                    f"instrument_model={instrument_model!r}",
                    enabled=True,
                )

            instrument_model = normalize_instrument_model(
                instrument_model if instrument_model != "unknown" else "",
                platform,
                allowed_instruments,
            )
            if debug and instrument_model == "":
                log_debug(
                    "SRA instrument filtered: "
                    f"instrument_model not in template list",
                    enabled=True,
                )

            seq_tech = (
                primary_reads.get("source_sequencing_technology")
                or primary_reads.get("sequencing_technology")
            )
            tech_label = infer_read_tech_label(platform, seq_tech)
            long_read = is_long_read_tech(platform, seq_tech)
            if (
                plasmidsaurus_year
                and prefer_plasmidsaurus_year_long_reads
                and long_read
                and reads_date
                and not reads_date.startswith(f"{plasmidsaurus_year}-")
            ):
                if debug:
                    log_debug(
                        "Skipping long read due to plasmidsaurus protocol year mismatch: "
                        f"isolate={isolate_name!r} reads={primary_reads.get('reads_name')!r} "
                        f"reads_date={reads_date!r} protocol_year={plasmidsaurus_year!r}",
                        enabled=True,
                    )
                continue

            if long_read:
                if tech_label == "Pacbio":
                    base_library_id = f"{isolate_name}_pacbio"
                else:
                    base_library_id = f"{isolate_name}_nano"
            else:
                base_library_id = isolate_name

            # Force ONT submissions to single-end metadata regardless of BERDL read_type.
            if platform == "OXFORD_NANOPORE":
                library_layout = "SINGLE"
                filename2 = None

            library_stem = base_library_id
            if (
                not long_read
                and short_read_date_flag.get(isolate_name)
                and reads_date
            ):
                library_stem = f"{base_library_id}_{reads_date}"
            elif library_stem in used_library_ids and reads_date:
                library_stem = f"{base_library_id}_{reads_date}"

            title_base = f"{tech_label} reads for {isolate_name}"
            if (
                not long_read
                and short_read_date_flag.get(isolate_name)
                and reads_date
            ):
                title_base = f"{title_base}, {reads_date}"
            elif library_stem.endswith(f"_{reads_date}") and reads_date:
                title_base = f"{title_base}, {reads_date}"

            replicate_from_filename = infer_replicate_number_from_filenames(
                filename, filename2, primary_reads.get("reads_name")
            )
            suffix_number: Optional[int] = None
            if replicate_from_filename is not None:
                suffix_number = replicate_from_filename
                library_id = f"{library_stem}_{suffix_number}"
                while library_id in used_library_ids:
                    suffix_number += 1
                    library_id = f"{library_stem}_{suffix_number}"
                next_suffix_by_library_stem[library_stem] = max(
                    next_suffix_by_library_stem.get(library_stem, 2), suffix_number + 1
                )
            else:
                library_id = library_stem
                if library_id in used_library_ids:
                    suffix_number = next_suffix_by_library_stem.get(library_stem, 2)
                    while f"{library_stem}_{suffix_number}" in used_library_ids:
                        suffix_number += 1
                    library_id = f"{library_stem}_{suffix_number}"
                    next_suffix_by_library_stem[library_stem] = suffix_number + 1
                else:
                    next_suffix_by_library_stem.setdefault(library_stem, 2)

            used_library_ids.add(library_id)

            title = title_base if suffix_number is None else f"{title_base}, {suffix_number}"
            design_description = infer_sra_design_description(
                platform,
                instrument_model,
                list(protocol_names) + list(assembly_protocols),
            )

            values = {
                "sample_name": biosample_name,
                "library_ID": library_id,
                "title": title,
                "library_strategy": "WGS",
                "library_source": "GENOMIC",
                "library_selection": "RANDOM",
                "library_layout": library_layout,
                "platform": platform,
                "instrument_model": instrument_model,
                "design_description": design_description,
                "filetype": "fastq",
                "filename": filename,
            }
            if filename2:
                values["filename2"] = filename2

            for header, value in values.items():
                col = header_map.get(header)
                if col is not None:
                    sheet.cell(row=next_row, column=col, value=value)
            next_row += 1

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not dry_run:
        sheet.parent.save(output_path)
    return max(0, next_row - start_row)


def generate_genome_table(
    genome_data: List[Dict[str, Any]],
    output_file: str,
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
    edr_root: str,
    debug: bool = False,
    dry_run: bool = False,
) -> int:
    template_path, sheet, header_row, header_map = load_genome_template_workbook(
        Path(output_file).parent
    )

    last_data_row = header_row
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        if any(cell.value not in (None, "") for cell in row):
            last_data_row = row[0].row

    next_row = last_data_row + 1
    start_row = next_row
    for genome in genome_data:
        strain = genome.get("strain", {}) or {}
        sample = genome.get("sample", {}) or {}
        raw_sample_name = sample.get("sample_name") or ""
        isolate_name = strain.get("strain_name") or raw_sample_name or genome.get(
            "genome_name", ""
        )
        if not raw_sample_name:
            raw_sample_name = isolate_name
        biosample_name = build_biosample_name(raw_sample_name, isolate_name)

        method_name, method_version, assembly_protocols = resolve_assembly_method_for_genome(
            genome,
            headers,
            column_cache,
            protocol_cache,
            edr_root=edr_root,
            debug=debug,
        )
        genome_name = genome.get("genome_name", "")

        seq_techs = infer_sequencing_techs(
            assembly_protocols, headers, column_cache, protocol_cache
        )
        if not seq_techs:
            for reads in genome.get("reads", []) or []:
                tech = reads.get("sequencing_technology")
                if tech:
                    seq_techs.append(tech)
        sequencing_technology = "; ".join(sorted(set(seq_techs))) if seq_techs else ""
        if (
            (not sequencing_technology)
            or sequencing_technology.strip().lower() in {"unknown", "none", "na", "n/a"}
        ) and str(method_name or "").strip().lower() == "spades":
            sequencing_technology = "Illumina"

        fasta_link, _ = select_contig_link_and_path(
            genome, edr_root=edr_root, debug=debug
        )
        fasta_filename = Path(fasta_link).name if fasta_link else ""
        assembly_date = choose_assembly_date(genome.get("assembly_processes", []) or [])

        values = {
            "biosample_accession": "",
            "sample_name": biosample_name,
            "assembly_date": assembly_date,
            "assembly_name": genome_name,
            "assembly_method": method_name,
            "assembly_method_version": method_version,
            "genome_coverage": genome.get("genome_coverage", ""),
            "sequencing_technology": sequencing_technology,
            "reference_genome": "",
            "update_for": genome.get("refseq_id", ""),
            "bacteria_available_from": Bacteria_AVAILABLE_FROM,
            "filename": fasta_filename,
        }

        for header, value in values.items():
            col = header_map.get(header)
            if col is not None:
                sheet.cell(row=next_row, column=col, value=value)
        next_row += 1

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not dry_run:
        sheet.parent.save(output_path)
    return max(0, next_row - start_row)


def _add_part_suffix(filename: str, part_index: int) -> str:
    path = Path(filename)
    return str(path.with_name(f"{path.stem}_part{part_index}{path.suffix}"))


def resolve_assembly_method_for_genome(
    genome: Dict[str, Any],
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
    edr_root: Optional[str] = None,
    debug: bool = False,
) -> Tuple[str, str, List[str]]:
    assembly_processes = genome.get("assembly_processes", []) or []
    assembly_protocols = collect_protocols_from_processes(assembly_processes)
    assembly_method = infer_assembly_method_from_protocols(
        assembly_protocols, headers, column_cache, protocol_cache
    )
    use_long_read_hardcoded_metadata = is_long_read_submission_context(
        assembly_protocols,
        genome.get("reads", []) or [],
        headers,
        column_cache,
        protocol_cache,
    )
    method_name, method_version = split_assembly_method(assembly_method, debug=debug)
    genome_name = genome.get("genome_name", "")
    strain_id = str(genome_name).split(".", 1)[0] if genome_name else ""
    submission_genome_name = str(genome_name or "")
    if submission_genome_name.endswith(".genome"):
        submission_genome_name = submission_genome_name[: -len(".genome")]
    has_lui_nielsen_2022_flye_unicycler_protocol = any(
        "lui-nielsen-2022-flye-unicycler" in str(protocol_name).lower()
        for protocol_name in assembly_protocols
    )
    has_price_2024_hifi_protocol = any(
        "price-2024-hifi" in str(protocol_name).lower()
        for protocol_name in assembly_protocols
    )
    has_lui_2020_spades_protocol = any(
        "lui-2020-spades" in str(protocol_name).lower()
        for protocol_name in assembly_protocols
    )
    strain = genome.get("strain", {}) or {}
    sample = genome.get("sample", {}) or {}
    raw_sample_name = sample.get("sample_name") or ""
    isolate_name = (
        strain.get("strain_name") or raw_sample_name or genome.get("genome_name", "")
    )
    hardcoded_meta = (
        hardcoded_metadata_for_isolate(isolate_name)
        if use_long_read_hardcoded_metadata
        else {}
    )
    hardcoded_assembler = hardcoded_meta.get("assembler", "")
    if use_long_read_hardcoded_metadata:
        enforce_assembler_match(strain_id, method_name, hardcoded_assembler)

    if use_long_read_hardcoded_metadata and not _is_metadata_unknown(hardcoded_assembler):
        method_name = normalize_hardcoded_assembler_name(hardcoded_assembler)
        if method_version in {"", "unknown"}:
            method_version = ""
    elif strain_id in FLYE_292_B1786_GENOMES:
        method_name = "Flye"
        method_version = "2.9.2-b1786"
    elif strain_id in FLYE_29_B1768_GENOMES:
        method_name = "Flye"
        method_version = "2.9-b1768"
    plasmidsaurus_protocol = False
    for protocol_name in assembly_protocols:
        info = get_protocol_info(headers, protocol_name, column_cache, protocol_cache)
        proto_text = " ".join(
            [
                str(info.get("protocol_name") or protocol_name),
                str(info.get("protocol_description") or ""),
            ]
        ).lower()
        if "plasmidsaurus" in proto_text:
            plasmidsaurus_protocol = True
            break
    if plasmidsaurus_protocol:
        method_name = "Flye"
        method_version = "2.9.6"
    if has_lui_2020_spades_protocol:
        method_name = "Unicycler"
        method_version = "0.4.8"
    if (
        submission_genome_name in LUI_NIELSEN_2022_FLYE_UNICYCLER_GENOMES
        and has_lui_nielsen_2022_flye_unicycler_protocol
    ):
        method_name = "Unicycler"
        method_version = "0.4.8"
    if (
        submission_genome_name in PRICE_2024_HIFI_GENOMES
        and has_price_2024_hifi_protocol
    ):
        method_name = "Flye"
        method_version = "2.9.2-b1786"
    if not method_name:
        method_name = "unknown"
    if not method_version and method_name == "unknown":
        method_version = "unknown"

    if method_name in {"Flye", "Unicycler"} and not method_version and edr_root:
        _, fasta_path = select_contig_link_and_path(genome, edr_root=edr_root, debug=debug)
        method_version = infer_assembler_version_from_edr_logs(
            fasta_path, method_name, debug=debug
        )
    method_version = normalize_assembler_version(method_version)
    return method_name, method_version, assembly_protocols


def write_remaining_unknown_assemblies_report(
    genome_data: List[Dict[str, Any]],
    output_file: str,
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
    edr_root: str,
    debug: bool = False,
) -> None:
    grouped: Dict[str, List[Tuple[str, str, str]]] = defaultdict(list)
    total = 0
    for genome in genome_data:
        method_name, method_version, assembly_protocols = resolve_assembly_method_for_genome(
            genome,
            headers,
            column_cache,
            protocol_cache,
            edr_root=edr_root,
            debug=debug,
        )
        unresolved = (
            (not method_name)
            or method_name.lower() == "unknown"
            or (not method_version)
            or method_version.lower() == "unknown"
        )
        if not unresolved:
            continue
        total += 1
        genome_name = str(genome.get("genome_name") or "")
        if not assembly_protocols:
            protocol_label = "<no assembly-like process found>"
        elif len(assembly_protocols) == 1:
            protocol_label = f"{assembly_protocols[0]} [Shotgun Sequencing and Assembly]"
        else:
            protocol_label = (
                f"{', '.join(assembly_protocols)} [Shotgun Sequencing and Assembly]"
            )
        version_display = method_version if method_version else "<blank>"
        grouped[protocol_label].append((genome_name, method_name or "unknown", version_display))

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append("Remaining genomes with unknown assembly method or version")
    lines.append(f"Total genomes: {total}")
    lines.append("")
    protocol_labels = sorted(label for label in grouped.keys() if label != "<no assembly-like process found>")
    if "<no assembly-like process found>" in grouped:
        protocol_labels = ["<no assembly-like process found>"] + protocol_labels
    for label in protocol_labels:
        rows = sorted(grouped[label], key=lambda row: row[0])
        lines.append(f"Protocol: {label}")
        lines.append(f"Count: {len(rows)}")
        for genome_name, method_name, method_version in rows:
            lines.append(
                f"- {genome_name}\tassembly_method={method_name}\tassembly_method_version={method_version}"
            )
        lines.append("")
    output_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def partition_genome_data_for_submission(
    genome_data: List[Dict[str, Any]],
    output_dir: str,
    headers: Dict[str, str],
    column_cache: Dict[str, List[str]],
    protocol_cache: Dict[str, Dict[str, Optional[str]]],
    max_genomes_per_file: int,
    max_sra_rows_per_file: int,
    debug: bool = False,
) -> List[List[Dict[str, Any]]]:
    if not genome_data:
        return []

    pending: List[List[Dict[str, Any]]] = [list(genome_data)]
    accepted: List[List[Dict[str, Any]]] = []

    while pending:
        part = pending.pop(0)
        biosample_rows = len(part)
        genome_rows = len(part)
        sra_rows = generate_sra_table(
            part,
            os.path.join(output_dir, "_dryrun_sra_table_SRA_metadata.xlsx"),
            headers,
            column_cache,
            protocol_cache,
            debug=debug,
            dry_run=True,
        )

        over_limit = (
            biosample_rows > max_genomes_per_file
            or genome_rows > max_genomes_per_file
            or sra_rows > max_sra_rows_per_file
        )
        if not over_limit:
            accepted.append(part)
            continue

        if len(part) <= 1:
            log_info(
                "WARNING: A single-genome part exceeds submission limits "
                f"({biosample_rows=}, {genome_rows=}, {sra_rows=}). "
                "Keeping this part unsplit."
            )
            accepted.append(part)
            continue

        midpoint = len(part) // 2
        pending.insert(0, part[midpoint:])
        pending.insert(0, part[:midpoint])

    return accepted


def process_genomes_for_submission(
    headers: Dict[str, str],
    genome_names: Sequence[str],
    output_dir: str = DEFAULT_OUTPUT_DIR,
    debug: bool = False,
    edr_path: str = DEFAULT_EDR_PATH,
    skip_coverage_calculation: bool = False,
    sample_metadata_path: Optional[str] = None,
) -> List[Dict[str, Any]]:
    if debug:
        set_debug(True)
        log_info("Debug enabled; BERDL API calls will be logged")
    os.makedirs(output_dir, exist_ok=True)
    column_cache: Dict[str, List[str]] = {}
    protocol_cache: Dict[str, Dict[str, Optional[str]]] = {}
    read_cache: Dict[str, Dict[str, Any]] = {}
    sample_metadata_map = load_sample_metadata(sample_metadata_path, debug=debug)
    total_metadata_fields_compared = 0
    total_metadata_fields_matched = 0
    total_samples_with_metadata = 0

    log_info(f"Preparing BERDL provenance data for {len(genome_names)} genome(s)")
    log_info("Discovering tables in BERDL")
    discovered_tables = discover_tables(headers)
    log_info(f"Discovered {len(discovered_tables)} BERDL tables")
    log_info("Loading sys_process cache (may take a while)")
    cache = load_process_cache(headers, discovered_tables)
    log_info(f"Loaded {len(cache.process_rows)} sys_process rows")
    downstream_lookup = build_downstream_lookup(cache.out_lookup)
    resolver = NameResolver(headers) if debug else None

    log_info("Fetching genome table schema")
    genome_columns = get_table_columns(headers, "sdt_genome", column_cache)
    genome_desired = [
        col
        for col in ["sdt_genome_id", "sdt_genome_name", "sdt_strain_name", "link"]
        if col in genome_columns
    ]

    genome_data: List[Dict[str, Any]] = []
    warnings: List[str] = []

    for idx, genome_name in enumerate(genome_names, start=1):
        log_info(f"[{idx}/{len(genome_names)}] Processing genome {genome_name}")
        genome_row = select_first_row(
            headers,
            "sdt_genome",
            [{"column": "sdt_genome_name", "operator": "=", "value": genome_name}],
            genome_desired,
        )
        if not genome_row:
            warnings.append(f"Genome {genome_name} not found in sdt_genome table")
            continue

        genome_id = genome_row.get("sdt_genome_id")
        if not genome_id:
            warnings.append(f"Genome {genome_name} missing sdt_genome_id")
            continue

        genome_token = f"sdt_genome:{genome_id}"
        strain_name = genome_row.get("sdt_strain_name")
        genome_link = genome_row.get("link")
        log_debug(
            f"Genome token={genome_token} strain={strain_name} link={genome_link}",
            enabled=debug,
        )

        if debug:
            log_info(f"Provenance for {genome_name} ({genome_token})")
            walk_provenance(genome_token, cache.out_lookup, resolver)  # type: ignore[arg-type]
        log_info(f"Finding reads for {genome_name}")
        strain_token = None
        if strain_name:
            strain_id = get_strain_id(headers, strain_name, column_cache)
            if strain_id:
                strain_token = f"sdt_strain:{strain_id}"
        reads_list = find_oldest_reads_with_fastq(
            genome_token,
            cache.out_lookup,
            downstream_lookup,
            headers,
            column_cache,
            read_cache,
            strain_token=strain_token,
            log_label=f"[reads {genome_name}] ",
            genome_name=genome_name,
        )
        if not reads_list:
            warnings.append(
                f"Genome {genome_name}: No reads with FASTQ files on {FASTQ_HOST} found"
            )
        log_info(f"Found {len(reads_list)} read set(s) with FASTQ for {genome_name}")

        log_info(f"Finding samples for {genome_name}")
        samples_list = find_samples_from_genome(
            genome_token, cache.out_lookup, headers, column_cache
        )
        log_info(f"Found {len(samples_list)} sample(s) for {genome_name}")
        sample_data = samples_list[0] if samples_list else None
        sample_metadata = {}
        if sample_data:
            sample_id = _normalize_text(sample_data.get("sample_name"))
            sample_metadata = sample_metadata_map.get(sample_id, {})
            if not sample_metadata and sample_metadata_map:
                warnings.append(
                    f"{genome_name}: sample {sample_id!r} not found in sample_metadata.tsv"
                )
        location_data = None
        if sample_data and sample_data.get("location_name"):
            log_info(f"Resolving location {sample_data['location_name']}")
            location_data = get_location_info(
                headers, sample_data["location_name"], column_cache
            )
        if sample_data and sample_metadata:
            match_summary = summarize_sample_metadata_match(
                sample_data, location_data or {}, sample_metadata
            )
            total_samples_with_metadata += 1
            total_metadata_fields_compared += int(match_summary["compared_count"])
            total_metadata_fields_matched += int(match_summary["matched_count"])
            if debug:
                log_debug(
                    f"Sample metadata match summary for {genome_name}: "
                    f"{match_summary['matched_count']}/{match_summary['compared_count']} matched; "
                    f"matched={match_summary['matched_fields']}; "
                    f"mismatched={match_summary['mismatched_fields']}",
                    enabled=True,
                )
            warnings.extend(
                validate_sample_metadata_match(
                    genome_name, sample_data, location_data or {}, sample_metadata
                )
            )

        if strain_name:
            log_info(f"Resolving strain {strain_name}")
        strain_data = get_strain_info(headers, strain_name, column_cache) if strain_name else None
        gtdb_genus = get_gtdb_genus_for_strain(headers, strain_name, column_cache)
        isolate_key = normalize_isolate_key(strain_name or genome_name)
        if isolate_key in HARDCODED_GTDB_GENUS_BY_ISOLATE:
            gtdb_genus = HARDCODED_GTDB_GENUS_BY_ISOLATE[isolate_key]
        collected_by = None
        if sample_data and sample_data.get("sample_token"):
            collected_by = resolve_collected_by(sample_data["sample_token"], cache.out_lookup)
        if collected_by is None:
            collected_by = resolve_collected_by(genome_token, cache.out_lookup)

        log_info(f"Finding assembly processes for {genome_name}")
        assembly_processes = find_assembly_processes(genome_token, cache.out_lookup)
        log_info(f"Found {len(assembly_processes)} assembly process(es) for {genome_name}")

        genome_data.append(
            {
                "genome_name": genome_name,
                "genome_id": genome_id,
                "genome_link": genome_link,
                "strain": strain_data or {"strain_name": strain_name},
                "sample": sample_data,
                "location": location_data,
                "reads": reads_list,
                "assembly_processes": assembly_processes,
                "gtdb_genus": gtdb_genus,
                "collected_by": collected_by,
                "sample_metadata": sample_metadata,
            }
        )

    if warnings:
        log_info("Warnings:")
        for warning in warnings:
            print(f"  - {warning}", file=sys.stderr)
    if sample_metadata_map:
        log_info(
            "Sample metadata overall match summary: "
            f"{total_metadata_fields_matched}/{total_metadata_fields_compared} matched "
            f"across {total_samples_with_metadata} sample(s) with metadata."
        )

    log_info("Fetching read coverage data")
    strain_names = [
        normalize_strain_name(genome.get("genome_name", ""), genome.get("strain"))
        for genome in genome_data
    ]
    strain_names = [name for name in strain_names if name]
    log_info("Fetching current isolate RefSeq IDs from BERDL")
    refseq_map = fetch_isolate_refseq_map(headers, strain_names, discovered_tables, column_cache)
    for genome in genome_data:
        strain_name = normalize_strain_name(genome.get("genome_name", ""), genome.get("strain"))
        refseq_id = refseq_map.get(strain_name) or refseq_map.get(
            normalize_isolate_key(strain_name)
        )
        if refseq_id:
            genome["refseq_id"] = refseq_id

    coverage_map = fetch_read_coverage_map(headers, strain_names, column_cache)
    if skip_coverage_calculation:
        log_info(
            "Skipping EDR file-based coverage fallback; using BERDL read coverage values only."
        )
    for genome in genome_data:
        strain_name = normalize_strain_name(genome.get("genome_name", ""), genome.get("strain"))
        if strain_name in coverage_map:
            genome["genome_coverage"] = coverage_map[strain_name]
        elif not skip_coverage_calculation:
            coverage = compute_coverage_from_files(genome, edr_path, debug=debug)
            if coverage is not None:
                genome["genome_coverage"] = coverage

    partitioned_genomes = partition_genome_data_for_submission(
        genome_data,
        output_dir,
        headers,
        column_cache,
        protocol_cache,
        max_genomes_per_file=MAX_GENOMES_PER_SUBMISSION,
        max_sra_rows_per_file=MAX_SRA_ROWS_PER_SUBMISSION,
        debug=debug,
    )
    split_mode = len(partitioned_genomes) > 1
    if split_mode:
        log_info(
            f"Splitting submission into {len(partitioned_genomes)} part(s) "
            f"to keep genome/biosample tables <= {MAX_GENOMES_PER_SUBMISSION} rows "
            f"and SRA table <= {MAX_SRA_ROWS_PER_SUBMISSION} rows."
        )

    generated_files: List[str] = []
    for idx, part_genomes in enumerate(partitioned_genomes, start=1):
        biosample_file = os.path.join(output_dir, "biosample_table_Microbe.1.0.xlsx")
        sra_file = os.path.join(output_dir, "sra_table_SRA_metadata.xlsx")
        genome_file = os.path.join(output_dir, "genome_table_Template_GenomeBatch.xlsx")
        reads_dir = "reads_to_upload"
        contigs_dir = "contigs_to_upload"
        if split_mode:
            biosample_file = _add_part_suffix(biosample_file, idx)
            sra_file = _add_part_suffix(sra_file, idx)
            genome_file = _add_part_suffix(genome_file, idx)
            reads_dir = f"reads_to_upload_part{idx}"
            contigs_dir = f"contigs_to_upload_part{idx}"

        if split_mode:
            log_info(
                f"Writing part {idx}/{len(partitioned_genomes)} "
                f"({len(part_genomes)} genome(s))"
            )
        log_info("Writing biosample table")
        generate_biosample_table(
            part_genomes,
            biosample_file,
            output_dir,
            debug=debug,
        )
        log_info("Writing SRA metadata table")
        generate_sra_table(
            part_genomes,
            sra_file,
            headers,
            column_cache,
            protocol_cache,
            debug=debug,
        )
        log_info("Writing genome metadata table")
        generate_genome_table(
            part_genomes,
            genome_file,
            headers,
            column_cache,
            protocol_cache,
            edr_root=edr_path,
            debug=debug,
        )
        log_info("Linking FASTQ files for upload")
        link_fastq_files(
            part_genomes,
            output_dir,
            edr_path,
            target_subdir=reads_dir,
            debug=debug,
        )
        log_info("Linking contig files for upload")
        link_contig_files(
            part_genomes,
            output_dir,
            edr_path,
            target_subdir=contigs_dir,
            debug=debug,
        )

        generated_files.extend([biosample_file, sra_file, genome_file])

    remaining_unknown_path = str(DEFAULT_REMAINING_UNKNOWN_ASSEMBLIES_PATH)
    log_info("Writing remaining unknown assemblies report")
    write_remaining_unknown_assemblies_report(
        genome_data,
        remaining_unknown_path,
        headers,
        column_cache,
        protocol_cache,
        edr_root=edr_path,
        debug=debug,
    )

    print("Generated submission tables:")
    for file_path in generated_files:
        print(f"  - {file_path}")
    print(f"  - {remaining_unknown_path}")

    return genome_data


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate NCBI submission tables using BERDL provenance data."
    )
    parser.add_argument(
        "--genome-name",
        action="append",
        dest="genome_names",
        help="Genome name from sdt_genome.sdt_genome_name (repeatable).",
    )
    parser.add_argument(
        "--genome-list",
        help="Path to a file with one genome name per line.",
    )
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT_DIR,
        help=f"Directory to write output files (default: {DEFAULT_OUTPUT_DIR}).",
    )
    parser.add_argument(
        "--base-url",
        default=os.environ.get("BERDL_BASE_URL", DEFAULT_BASE_URL),
        help=f"MCP base URL (default: {DEFAULT_BASE_URL}).",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable verbose debugging, including BERDL API calls.",
    )
    parser.add_argument(
        "--edr-path",
        default=DEFAULT_EDR_PATH,
        help=f"Path to ENIGMA data repository root (default: {DEFAULT_EDR_PATH}).",
    )
    parser.add_argument(
        "--skip-coverage-calculation",
        action="store_true",
        help=(
            "Skip EDR file-based coverage fallback. "
            "Use only genome coverage values available in BERDL."
        ),
    )
    parser.add_argument(
        "--sample-metadata",
        default=None,
        help=(
            "Optional path to sample metadata TSV (default search order: "
            "sample_metadata.tsv, genome_upload/sample_metadata.tsv)."
        ),
    )
    return parser.parse_args()


def load_genome_names(args: argparse.Namespace) -> List[str]:
    names = list(args.genome_names or [])
    if args.genome_list:
        with open(args.genome_list, "r", encoding="utf-8") as handle:
            for line in handle:
                name = line.strip()
                if name:
                    names.append(name)
    deduped = []
    seen: set[str] = set()
    for name in names:
        if name in seen:
            continue
        seen.add(name)
        deduped.append(name)
    return deduped


def main() -> None:
    args = parse_args()
    walk_provenance_module.BASE_URL = args.base_url
    enable_request_failure_logging()
    genome_names = load_genome_names(args)
    if not genome_names:
        raise SystemExit("No genomes provided. Use --genome-name or --genome-list.")
    headers = get_headers()
    process_genomes_for_submission(
        headers,
        genome_names,
        output_dir=args.output_dir,
        debug=args.debug,
        edr_path=args.edr_path,
        skip_coverage_calculation=args.skip_coverage_calculation,
        sample_metadata_path=args.sample_metadata,
    )


if __name__ == "__main__":
    main()
